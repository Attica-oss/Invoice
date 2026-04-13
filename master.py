import marimo

__generated_with = "0.23.1"
app = marimo.App(width="medium")

with app.setup:
    import polars as pl
    import openpyxl
    import marimo as mo


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    # Transfer of Scows
    """)
    return


@app.cell
def _():
    from dataframe import full_scows,empty_scows

    return empty_scows, full_scows


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Full Scow Transfer

    *  Has the Trucking price of $3.5/tons,
    *  Has the CCCS movement fee of $3.5/tons
    """)
    return


@app.cell(hide_code=True)
def _(full_scows):
    full_scow_transfer_df = mo.sql(
        f"""
        FROM
            full_scows
        SELECT
            day_name,
            date,
            customer,
            CASE
                WHEN movement_type LIKE '%Delivery%' THEN 'Delivery'
                WHEN movement_type LIKE '%Collection%' THEN 'Collection'
                ELSE 'Invalid Move'
            END AS movement_type,
            overtime,
            num_of_scows AS number_of_scows,
            tonnage,
            storage_type,
            Price AS unit_price,
            total_price AS movement_fee,
            total_price AS scow_transfer_price,
            'FULL SCOW TRANSFER' AS service,
            'BIN DISPATCH TO/FROM IOT' AS sub_type,
        'MONTHLY' AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Empty Scow Transfer

    *  Has the Trucking price of $3.5/scows,
    """)
    return


@app.cell
def _(empty_scows):
    empty_scow_transfer_df = mo.sql(
        f"""
        FROM
            empty_scows
        SELECT
            day_name,
            date,
            customer,
            CASE
                WHEN movement_type LIKE '%Delivery%' THEN 'Delivery'
                WHEN movement_type LIKE '%Collection%' THEN 'Collection'
                ELSE 'Invalid Move'
            END AS movement_type,
            overtime,
            num_of_scows AS number_of_scows,
            -- tonnage,
            -- storage_type,
            Price AS unit_price,
            -- total_price AS movement_fee,
            total_price AS scow_transfer_price,
            'EMPTY SCOW TRANSFER' AS service,
            'MONTHLY' AS sub_type,
        'MONTHLY' AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Forklift Scow Handling

    * For both empty and full scow handling
    """)
    return


@app.cell(hide_code=True)
def _(empty_scows, full_scows):
    forklift_scow_handling = mo.sql(
        f"""
        WITH full_ AS (FROM
            full_scows
        SELECT
            day_name,
            date,
            customer,
            CASE
                WHEN movement_type LIKE '%Delivery%' THEN 'Loading Full Scows'
                WHEN movement_type LIKE '%Collection%' THEN 'Unloading Full Scows'
                ELSE 'Invalid Move'
            END AS forklift_service,
            overtime,
            num_of_scows AS number_of_scows,
            -- tonnage,
            -- storage_type,
            6 AS unit_price,
            CASE
                WHEN overtime = 'overtime 200%' THEN unit_price * num_of_scows * 2.0
                WHEN overtime = 'overtime 150%' THEN unit_price * num_of_scows * 1.5
                WHEN overtime = 'normal hours' THEN unit_price * num_of_scows * 1.0
                ELSE 0
            END AS forklift_price,
            'FORKLIFT SCOW HANDLING' AS service,
            'BIN DISPATCH TO/FROM IOT' AS sub_type,
            'MONTHLY' AS report_type), 
            empty_ AS (FROM
            empty_scows
        SELECT
            day_name,
            date,
            customer,
            CASE
                WHEN movement_type LIKE '%Delivery%' THEN 'Loading Empty Scows'
                WHEN movement_type LIKE '%Collection%' THEN 'Unloading Empty Scows'
                ELSE 'Invalid Move'
            END AS forklift_service,
            overtime,
            num_of_scows AS number_of_scows,
            -- tonnage,
            -- storage_type,
            6 AS unit_price,
            CASE
                WHEN overtime = 'overtime 200%' THEN unit_price * num_of_scows * 2.0
                WHEN overtime = 'overtime 150%' THEN unit_price * num_of_scows * 1.5
                WHEN overtime = 'normal hours' THEN unit_price * num_of_scows * 1.0
                ELSE 0
            END AS forklift_price,
            'FORKLIFT SCOW HANDLING' AS service,
            'MONTHLY' AS sub_type,
            'MONTHLY' AS report_type)

        FROM full_ 
        UNION ALL
        FROM empty_

        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## EMR

    * Washing
    * Pre-Trip Inspection
    """)
    return


@app.cell(hide_code=True)
def _():
    from dataframe import washing,pti,shifting

    return pti, shifting, washing


@app.cell(hide_code=True)
def _(washing):
    washing_df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    washing
                SELECT
                    *,
                    CASE
                        WHEN invoice_to IN ('MAERSKLINE', 'CMA CGM') THEN 'MONTHLY'
                        WHEN invoice_to = 'IOT' THEN 'M&R'
                        WHEN invoice_to IN ('INVALID', 'IPHS') THEN 'NOT TO INVOICE'
                        WHEN invoice_to NOT IN ('MAERSKLINE', 'CMA CGM', 'IOT')
                        AND service_remarks = 'Unstuffed' THEN 'CROSS STUFFING SI'
                        ELSE 'SI'
                    END AS sub_type
            ),
            grouped_data_ AS (
                FROM
                    data_
                SELECT
                    * EXCLUDE (price, container_number, service_remarks),
                    SUM(price) AS washing_price,
                    COUNT(*) AS number_of_service
                GROUP BY ALL
                ORDER BY
                    date
            )
        FROM
            grouped_data_
        SELECT
           STRFTIME(date,'%a') AS day_name,
            date,
            invoice_to AS customer,
            number_of_service,
            30 AS unit_price,
            washing_price,
            'CONTAINER WASHING' AS service,
            sub_type,
            CASE
                WHEN sub_type IN ('MONTHLY', 'M&R') THEN 'MONTHLY'
                WHEN sub_type = 'NOT TO INVOICE' THEN 'NOT TO INVOICE'
                ELSE 'SI'
            END AS report_type
        """
    )
    return


@app.cell
def _(pti):
    pti_df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    pti
                SELECT
                    STRFTIME(datetime_start, '%a') AS day_name,
                    datetime_start::DATE AS date,
                    CASE
                        WHEN set_point = -25 THEN 'STANDARD'
                        WHEN set_point = -35 THEN 'MAGNUM'
                        WHEN set_point = -60 THEN 'S FREEZER'
                        ELSE 'INVALID'
                    END AS container_type,
                    invoice_to AS customer,
                    CASE
                        WHEN "hours" > 8 THEN 'ABOVE'
                        ELSE 'BELOW'
                    END AS duration,
                    status,
                    CASE
                        WHEN no_shifting IS TRUE THEN 1
                        ELSE 0
                    END AS count_shifting,
                    plugin_price,
                    electricity_price,
                    shifting_price,
                    total_price,
                    'PTI' AS service,
                    CASE
                        WHEN customer = 'IOT' THEN 'M&R'
                        WHEN customer IN ('MAERSKLINE', 'CMA CGM') THEN 'MONTHLY'
                        ELSE 'NOT TO INVOICE'
                    END AS sub_type,
                    'MONTHLY' AS report_type
                WHERE
                    YEAR(datetime_start) = 2026
            )
        FROM
            data_
        SELECT
            day_name,
            date,
            container_type,
            customer,
            duration,
            status,
            COUNT(*) AS number_of_service,
            SUM(count_shifting) AS number_of_shifts,
            SUM(plugin_price) AS plugin_price,
            SUM(electricity_price) AS electricity_price,
            SUM(shifting_price) AS shifting_price,
            SUM(total_price) AS total_pti_price,
            service,
            sub_type,
            report_type
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Shifting
    """)
    return


@app.cell(hide_code=True)
def _(cross_stuffing, washing):
    _df = mo.sql(
        f"""
        WITH
            washing_ AS (
                FROM
                    washing
                SELECT
                    *
                WHERE
                    invoice_to NOT IN ('MAERSKLINE', 'CMA CGM', 'INVALID', 'IPHS')
            ),
            cross_ AS (
                FROM
                    cross_stuffing
                SELECT
                    origin AS container_number,
                    * EXCLUDE (origin)
                WHERE
                    is_origin_empty = TRUE
                    AND invoiced NOT IN ('MAERSKLINE', 'CMA CGM', 'INVALID', 'IPHS')
            ),
            combined_data AS (
                SELECT
                    c.day AS day_name,
                    w.date,
                    w.invoice_to AS customer,
                    'Shifting at Washing' AS operation_type,
                    'CROSS STUFFING' AS service,
                       CASE
                WHEN c.day IN ('Sun', 'PH') THEN 35 * 1.5
                ELSE 35
            END AS shifting_price,
                    'SI' AS sub_type,
                    'SI' AS report_type
                FROM
                    cross_ c
                    LEFT JOIN washing_ w ON c.date <= w.date
                    AND w.container_number = c.container_number
                WHERE
                    w.date IS NOT NULL
            )
        FROM
            combined_data
        SELECT
            day_name,
            date,
            customer,
            operation_type,
            COUNT(*) AS number_of_service,
        SUM(shifting_price) AS shifting_price,
            service,
            sub_type,
            report_type
        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _(shifting):
    internal_shifting_df = mo.sql(
        f"""
        WITH data_ AS (FROM
            shifting
        SELECT
            day_name,
            date,
            invoice_to AS customer,
            CASE
                WHEN CONTAINS(service_remarks, 'OT')
                AND day_name IN ('Sun', 'PH') THEN price * 2.0
                WHEN CONTAINS(service_remarks, 'OT')
                AND day_name NOT IN ('Sun', 'PH') THEN price * 1.5
                ELSE price
            END AS price,
            'INTERNAL SHIFTING' AS service,
            CASE
                WHEN invoice_to IN (
                    'MAERSKLINE',
                    'CMA CGM',
                    'IOT',
                    'CCCS',
                    'ISLAND CATCH'
                ) THEN 'MONTHLY'
                ELSE 'SI'
            END AS sub_type,
            sub_type AS report_type
        WHERE
            YEAR(date) = 2026)

        FROM data_ 
        SELECT day_name,date,customer,COUNT(*) AS number_of_service,SUM(price) AS shifting_price,service,sub_type,report_type
        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Cold Store Services
    """)
    return


@app.cell
def _():
    from dataframe import dispatch_to_cargo,from_cccs_to_vessel,cross_stuffing,by_catch,cccs_stuffing

    return (
        by_catch,
        cccs_stuffing,
        cross_stuffing,
        dispatch_to_cargo,
        from_cccs_to_vessel,
    )


@app.cell(hide_code=True)
def _(dispatch_to_cargo):
    _df = mo.sql(
        f"""
        FROM
            dispatch_to_cargo
        SELECT
            day_name,
            date,
            vessel,
            customer,
            operation_type,
            total_tonnage,
            overtime_tonnage,
            CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(
                    (
                        (total_tonnage - overtime_tonnage) * truck_price * 1.5
                    ) + (overtime_tonnage * truck_price * 2.0)
                 AS DECIMAL)
                ELSE CAST(
                    (
                        (total_tonnage - overtime_tonnage) * truck_price * 1.0
                    ) + (overtime_tonnage * truck_price * 1.5)
                AS DECIMAL)
            END AS truck_price,
            CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(
                    (
                        (total_tonnage - overtime_tonnage) * cccs_movement_fee * 1.5
                    ) + (overtime_tonnage * cccs_movement_fee * 2.0)
               AS DECIMAL )
                ELSE CAST(
                    (
                        (total_tonnage - overtime_tonnage) * cccs_movement_fee * 1.0
                    ) + (overtime_tonnage * cccs_movement_fee * 1.5)
                AS DECIMAL)
            END AS cccs_movement_fee,
            CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(
                    (
                        (total_tonnage - overtime_tonnage) * stevedores_on_cargo_fee * 1.5
                    ) + (overtime_tonnage * stevedores_on_cargo_fee * 2.0)
              AS DECIMAL  )
                ELSE CAST(
                    (
                        (total_tonnage - overtime_tonnage) * stevedores_on_cargo_fee * 1.0
                    ) + (overtime_tonnage * stevedores_on_cargo_fee * 1.5)
               AS DECIMAL )
            END AS stevedores_on_cargo_fee,
        CAST(total_price AS DECIMAL) AS total_dispatch_price,
        storage_type,
        'DISPATCH TO/FROM CARGO VESSEL' AS service,
        'SI' AS sub_type,
        'SI' AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _(from_cccs_to_vessel):
    _df = mo.sql(
        f"""
        FROM
            from_cccs_to_vessel
        SELECT
            "day" AS day_name,
            date,
            vessel,
            customer,
            'From Cold Store To Vessel' AS operation_type,
            total_tonnage,
            overtime_tonnage,
        'Dry' AS storage_type,
         CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(
                    (
                        (total_tonnage - overtime_tonnage) * 7.5 * 1.5
                    ) + (overtime_tonnage * 7.5 * 2.0)
                 AS DECIMAL)
                ELSE CAST(
                    (
                        (total_tonnage - overtime_tonnage) * 7.5 * 1.0
                    ) + (overtime_tonnage * 7.5 * 1.5)
                AS DECIMAL)
            END AS truck_price,
            CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(
                    (
                        (total_tonnage - overtime_tonnage) * 3.5 * 1.5
                    ) + (overtime_tonnage * 3.5 * 2.0)
               AS DECIMAL )
                ELSE CAST(
                    (
                        (total_tonnage - overtime_tonnage) * 3.5 * 1.0
                    ) + (overtime_tonnage * 3.5 * 1.5)
                AS DECIMAL)
            END AS cccs_movement_fee,
        total_price,
        'DISPATCH TO/FROM VESSEL' AS service,
        'SI' AS sub_type,
        'SI' AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _(cross_stuffing):
    _df = mo.sql(
        f"""
        FROM
            cross_stuffing
        SELECT
            "day" AS day_name,
            date,
            vessel_client AS vessel,
            invoiced AS customer,
            Service AS operation_type,
            total_tonnage,
            overtime_tonnage,
            Price AS unit_price,
            total_price,
        'CROSS STUFFING' AS service,
        CASE 
        	WHEN invoiced IN ('MAERSKLINE','IOT','CMA CGM') THEN 'MONTHLY' WHEN invoiced IN ('INVALID','IPHS') THEN 'NOT TO INVOICE' ELSE 'SI' END AS sub_type,
        sub_type AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Extra Cross Stuffing Services

    * Tally
    * Tare Rental
    """)
    return


@app.cell
def _():
    from scan_google_sheet import scan_google_sheet

    return (scan_google_sheet,)


@app.cell
def _():
    misc_url = "https://docs.google.com/spreadsheets/d/1VbfiiWsp8yxs6KSR1CXpw1S_35tYlWV8UjjWah9Afpw/edit?pli=1&gid=757201340#gid=757201340"

    cross_stuffing_tab= "CrossStuffing"
    return cross_stuffing_tab, misc_url


@app.cell
def _(cross_stuffing_tab, misc_url, scan_google_sheet):
    extra_cross_stuffing_service = scan_google_sheet(url=misc_url,sheet_name=cross_stuffing_tab).filter(pl.col("date").dt.year().eq(2026))
    return (extra_cross_stuffing_service,)


@app.cell(hide_code=True)
def _(extra_cross_stuffing_service):
    _df = mo.sql(
        f"""
        FROM extra_cross_stuffing_service
        SELECT "day" AS day_name,date,vessel_client AS vessel,invoiced AS customer,total_tonnage,overtime_tonnage,tally
        WHERE service <> 'Unstuffing to CCCS' AND tally <> 'NA'
        """
    )
    return


@app.cell(hide_code=True)
def _(by_catch):
    _df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    by_catch
                SELECT
                    "day" AS day_name,
                    date,
                    vessel,
                    customer,
                    operation_type,
                    total_tonnage,
                    overtime_tonnage,
                    Price AS unit_price,
                    CAST(total_price AS DECIMAL) AS total_price,
                    storage_type,
                    CASE
                        WHEN operation_type = 'CCCS (By-Catch)' THEN 'BY CATCH HANDLING'
                        ELSE 'TRANSFER OF BY CATCH'
                    END AS service,
                    'MONTHLY' AS sub_type,
                    'MONTHLY' AS report_type
            )
        FROM
            data_
        SELECT
            day_name,
            date,
            customer,
            COUNT(*) AS number_of_service,
            CAST(SUM(total_tonnage) AS DECIMAL) AS total_tonnage,
            SUM(overtime_tonnage) AS overtime_tonnage,
            unit_price,
            SUM(total_price) AS total_price,
            service,
            sub_type,
            report_type
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell(hide_code=True)
def _(cccs_stuffing):
    _df = mo.sql(
        f"""
        FROM
            cccs_stuffing
        SELECT
            day_name,
            date,
            customer AS vessel,
            invoiced AS customer,
            COUNT(*) AS number_of_container,
            Service AS operation_type,
            CAST(SUM(total_tonnage) AS DECIMAL) AS total_tonnage,
            SUM(overtime_tonnage) AS overtime_tonnage,
            Price AS unit_price,
            storage_type,
            CAST(SUM(total_price) AS DECIMAL) AS total_price,
            'COLD STORE STUFFING' AS service,
            CASE WHEN invoiced = 'MAERSKLINE' THEN 'CCCS OSS' ELSE 'SI' END AS sub_type,
            'OSS' AS report_type
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell
def _():
    from dataframe import oss,netList,iot_stuffing

    return iot_stuffing, oss


@app.cell(hide_code=True)
def _(iot_stuffing):
    _df = mo.sql(
        f"""
        WITH data_ AS (FROM
            iot_stuffing
        SELECT
            day_name,
            date,
            vessel,
            'IOT' AS customer,
            overtime,
            total_tonnage,
            Price AS unit_price,
            "storage" AS storage_type,
            CAST(total_price AS DECIMAL) AS total_price) 

        FROM data_ 
        SELECT day_name,date,vessel,customer,overtime,count(*) AS number_of_containers,CAST(SUM(total_tonnage) AS DECIMAL) AS total_tonnage,unit_price,SUM(total_price) AS total_price,
        'CONTAINER STUFFING' AS service,'MONTHLY' AS sub_type,'MONTHLY' AS report_type

        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _(oss):
    _df = mo.sql(
        f"""
        FROM oss
        SELECT 
         STRFTIME(date,'%a') AS day_name,
        date,
        vessel,
            'MAERSKLINE' AS customer,
            COUNT(*) AS number_of_containers,
        overtime,
            storage_type,
        CAST(SUM(total_tonnage) AS DECIMAL) AS total_tonnage,
        service AS operation_type,
        Price AS unit_price,
        CAST(SUM(invoice_value) AS DECIMAL) AS total_price,
            'CONTAINER STUFFING' AS service,
        UPPER(operation_type) AS sub_type,
        'OSS' AS report_type,
        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _(netlist):
    _df = mo.sql(
        f"""
        FROM netList
            SELECT STRFTIME(date,'%a') AS day_name,
            date,
            CASE WHEN remarks IS NULL THEN vessel ELSE remarks || ' // EX-' ||vessel END AS vessel,
            service AS operation_type,
            overtime,
            CAST(SUM(total_tonnage) AS DECIMAL) AS total_tonnage,
            storage_type,
            Price AS unit_price,
            CAST(SUM(invoice_value) AS DECIMAL) AS total_price,
            'UNLOADING VESSEL' AS service,
            'STO' AS sub_type,
            'STO' AS report_type
        WHERE YEAR(date) = 2026
        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell
def _():
    from dataframe import salt,forklift_for_salt

    return forklift_for_salt, salt


@app.cell(hide_code=True)
def _(salt):
    _df = mo.sql(
        f"""
        FROM
            salt
        SELECT
            day_name,
            date,
            vessel,
            customer,
            operation_type,
                CASE
                WHEN EXTRACT(minute FROM duration) > 0
                    THEN EXTRACT(hour FROM duration) + 1
                ELSE EXTRACT(hour FROM duration)
            END AS billed_hours,
            SUM(tonnage) AS total_tonnage,
            SUM(normal) AS normal_hour_tonnage,
            SUM(overtime_150) AS overtime_150_tonnage,
            SUM(overtime_200) AS overtime_200_tonnage,
            CAST(SUM(price) AS DECIMAL) AS total_price,
            'SALT OPERATION' AS service,
            CASE
                WHEN customer IN ('HARTSWATER LIMITED', 'ECHEBASTAR') THEN 'MONTHLY'
                ELSE 'STO'
            END AS sub_type,
            sub_type AS report_type
        WHERE
            YEAR(date) = 2026
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell(hide_code=True)
def _(forklift_for_salt):
    _df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    forklift_for_salt
                SELECT
                    "day" AS day_name,
                    date,
                    vessel,
                    customer,
                    CAST(
                        regexp_replace(purpose, '[^0-9]', '', 'g') AS FLOAT
                    ) AS quantity,
                    CASE
                        WHEN EXTRACT (
                            minute
                            FROM
                                total_duration
                        ) > 0 THEN EXTRACT (
                            hour
                            FROM
                                total_duration
                        ) + 1
                        ELSE EXTRACT (
                            hour
                            FROM
                                total_duration
                        )
                    END AS billed_hours,
                    CASE
                        WHEN EXTRACT (
                            minute
                            FROM
                                overtime_for_normal_services
                        ) > 0 THEN EXTRACT (
                            hour
                            FROM
                                overtime_for_normal_services
                        ) + 1
                        ELSE EXTRACT (
                            hour
                            FROM
                                overtime_for_normal_services
                        )
                    END AS overtime_for_normal_services_hours,
                    CASE
                        WHEN EXTRACT (
                            minute
                            FROM
                                overtime_for_extended_services
                        ) > 0 THEN EXTRACT (
                            hour
                            FROM
                                overtime_for_extended_services
                        ) + 1
                        ELSE EXTRACT (
                            hour
                            FROM
                                overtime_for_extended_services
                        )
                    END AS overtime_for_extended_services_hours,
                    CASE
                        WHEN EXTRACT (
                            minute
                            FROM
                                normal_hour_services
                        ) > 0 THEN EXTRACT (
                            hour
                            FROM
                                normal_hour_services
                        ) + 1
                        ELSE EXTRACT (
                            hour
                            FROM
                                normal_hour_services
                        )
                    END AS normal_hour_services_hours
                WHERE
                    YEAR(date) = 2026
            )
        FROM
            data_
        SELECT
            day_name,
            date,
            vessel,
            customer,
            quantity,
            normal_hour_services_hours + overtime_for_normal_services_hours + overtime_for_extended_services_hours AS billed_hours,
            (overtime_for_normal_services_hours * 45) + (overtime_for_extended_services_hours * 60) + (normal_hour_services_hours * 30) AS total_price,
        'FORKLIFT FOR SALT OPERATION' AS service,
        CASE
                WHEN customer IN ('HARTSWATER LIMITED', 'ECHEBASTAR') THEN 'MONTHLY'
                ELSE 'STO'
            END AS sub_type,
            sub_type AS report_type
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    # Electricity
    """)
    return


@app.cell
def _():
    from dataframe import coa

    return (coa,)


@app.cell(hide_code=True)
def _(coa):
    _df = mo.sql(
        f"""
        FROM coa
        """
    )
    return


@app.cell
def _(coa):
    staged_electricity = mo.sql(
        f"""
        WITH
            monthly_bounds AS (
                SELECT
                    month_start::DATE AS month_start,
                    (
                        month_start + INTERVAL '1 month' - INTERVAL '1 day'
                    )::DATE AS month_end
                FROM
                    (
                        SELECT
                            unnest(
                                generate_series(
                                    '2026-01-01'::DATE,
                                    '2026-12-01'::DATE,
                                    INTERVAL '1 month'
                                )
                            ) AS month_start
                    )
            ),
            source AS (
                SELECT
                    vessel_client,
                    customer,
                    container_number,
                    operation_type,
                    GREATEST(date_plugged, '2026-01-01'::DATE) AS effective_start,
                    date_out,
                    plugin_price,
                    monitoring_price,
                    electricity_unit_price
                FROM
                    coa
                WHERE
                    date_out IS NULL
                    OR YEAR(date_out) = 2026
            ),
            split AS (
                SELECT
                    s.vessel_client,
                    s.customer,
                    s.container_number,
                    b.month_start AS invoice_month,
                    s.operation_type,
                    GREATEST(s.effective_start, b.month_start) AS period_start,
                    LEAST(s.date_out, b.month_end) AS period_end,
                    s.plugin_price,
                    s.monitoring_price,
                    s.electricity_unit_price,
                    -- Flag first and last month per container
                    ROW_NUMBER() OVER (
                        PARTITION BY
                            s.container_number
                        ORDER BY
                            b.month_start ASC
                    ) AS rn_first,
                    ROW_NUMBER() OVER (
                        PARTITION BY
                            s.container_number
                        ORDER BY
                            b.month_start DESC
                    ) AS rn_last
                FROM
                    source s
                    JOIN monthly_bounds b ON s.effective_start <= b.month_end
                    AND s.date_out >= b.month_start
            ),
            final AS (
                SELECT
                    vessel_client,
                    customer,
                    container_number,
                    operation_type,
                    invoice_month,
                    period_start,
                    period_end,
                    (period_end - period_start) AS days_on_plug,
                    electricity_unit_price,
                    -- Plugin only on first month
                    CASE
                        WHEN rn_first = 1 THEN plugin_price
                        ELSE 0
                    END AS plugin_fee,
                    -- Monitoring only on last month
                    CASE
                        WHEN rn_last = 1 THEN monitoring_price
                        ELSE 0
                    END AS monitoring_fee,
                    -- Electricity every month, prorated by days
                    (period_end - period_start) * electricity_unit_price AS electricity_fee
                FROM
                    split
                ORDER BY
                    vessel_client,
                    container_number,
                    invoice_month
            ),
            add_total_price AS (
                FROM
                    final
                SELECT
                    *,
                    (plugin_fee + monitoring_fee + electricity_fee) AS total_price
            )
        FROM
            add_total_price
        ORDER BY
            period_start
        """
    )
    return (staged_electricity,)


@app.cell(hide_code=True)
def _(staged_electricity):
    _df = mo.sql(
        f"""
        FROM
            staged_electricity
        SELECT
            invoice_month,
            vessel_client AS vessel,
            customer,
            operation_type,
            SUM(days_on_plug) AS number_of_days,
            SUM(plugin_fee) AS plugin_price,
            SUM(monitoring_fee) AS monitoring_price,
            SUM(electricity_fee) AS electricity_price,
            SUM(total_price) AS total_price,
            'ELECTRICITY' AS service,
            CASE
                WHEN operation_type LIKE '%CCCS Stuffing - Basic OSS%' THEN 'CCCS OSS'
                WHEN operation_type LIKE '%CCCS%' THEN 'MONTHLY'

                WHEN customer = 'IOT' THEN 'MONTHLY'
                WHEN operation_type LIKE '%Full OSS%' THEN 'FULL OSS'
                WHEN operation_type LIKE '%Basic OSS%' THEN 'BASIC OSS'
                WHEN operation_type = 'Plugin' THEN 'SI'
                WHEN operation_type LIKE '%Cross Stuffing%'
                AND customer = 'MAERSKLINE' THEN 'CROSS STUFFING OSS'
                WHEN operation_type LIKE '%Cross Stuffing%' THEN 'SI'
                WHEN operation_type = 'Stuffing and Plugin' THEN 'STO'
               WHEN customer = 'MAERSKLINE' AND vessel = 'MAERSKLINE' THEN 'MONTHLY'
                ELSE 'ERROR'
            END AS sub_type,
            CASE
                WHEN operation_type LIKE '%OSS%'
                OR (
                    operation_type LIKE '%Cross Stuffing%'
                    AND customer = 'MAERSKLINE'
                ) THEN 'OSS'
                WHEN customer = 'IOT' THEN 'MONTHLY'
                WHEN operation_type LIKE '%CCCS%' THEN 'MONTHLY'
                WHEN operation_type LIKE '%Cross Stuffing%'
                OR operation_type = 'Plugin' THEN 'SI'
                WHEN operation_type = 'Stuffing and Plugin' THEN 'STO'
            WHEN customer = 'MAERSKLINE' AND vessel = 'MAERSKLINE' THEN 'MONTHLY'
                ELSE 'ERROR'
            END AS report_type
        GROUP BY ALL
        HAVING
            number_of_days > 0
        ORDER BY
            invoice_month,
            vessel
        """
    )
    return


@app.cell
def _():
    from dataframe import pallet

    return (pallet,)


@app.cell
def _():
    exchange_rate_data = "./data/eur_usd_2026.csv"

    exchange_rate_df = pl.read_csv(exchange_rate_data)
    return (exchange_rate_df,)


@app.cell(hide_code=True)
def _(exchange_rate_df):
    _df = mo.sql(
        f"""
        FROM
            exchange_rate_df
        SELECT
            -- date::DATE AS date,
            month(date::DATE) AS month_number,
            AVG(eur_usd) AS exchange_rate
        GROUP BY month_number
        """
    )
    return


@app.cell(hide_code=True)
def _(exchange_rate_df, pallet):
    _df = mo.sql(
        f"""
        WITH
            pallet_ AS (
                FROM
                    pallet
                WHERE
                    YEAR(date) = 2026
            ),
            exc AS (
                FROM
                    exchange_rate_df
                SELECT
                    -- date::DATE AS date,
                    month(date::DATE) AS month_number,
                    AVG(eur_usd) AS exchange_rate
                GROUP BY
                    month_number
            ),
            keep_pallet_only AS (
                SELECT
                    p.*,
                    e.exchange_rate * p.pallet_price AS total_price
                FROM
                    pallet_ p
                    LEFT JOIN exc e ON MONTH(p.date) = e.month_number
                WHERE
                    remarks LIKE '%Pallet%'
            )
        FROM
            keep_pallet_only
        SELECT
            STRFTIME(date, '%a') AS day_name,
            date,
            assigned_to AS vessel,
            shipping_line AS customer,
            remarks AS operation_type,
            COUNT(*) AS number_of_service,
            CAST(SUM(total_price) AS DECIMAL) AS total_price,
            'PALLET INSTALLATION' AS service,
            CASE WHEN (customer = 'IOT') OR (customer IN ('AMIRANTE','ISLAND CATCH','OCEAN BASKET')) THEN 'MONTHLY' ELSE 'STO' END AS sub_type,
            sub_type AS report_type
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell(hide_code=True)
def _(pallet):
    _df = mo.sql(
        f"""
        WITH
            pallet_ AS (
                FROM
                    pallet
                WHERE
                    YEAR(date) = 2026

            ),
            keep_pallet_only AS (
                SELECT
                    p.* EXCLUDE(pallet_price,liner_price),
                    p.liner_price AS total_price
                FROM
                    pallet_ p
                WHERE
                    remarks LIKE '%Liner%'
            )
        FROM
            keep_pallet_only
        SELECT
            STRFTIME(date, '%a') AS day_name,
            date,
            assigned_to AS vessel,
            shipping_line AS customer,
            remarks AS operation_type,
        	COUNT(*) AS number_of_service,
        	CAST(SUM(total_price) AS DECIMAL) AS total_price,
        'LINER INSTALLATION' AS service,
            'MONTHLY' AS sub_type,
            'MONTHLY' AS report_type
        GROUP BY ALL
        ORDER BY date
        """
    )
    return


@app.cell
def _():
    from dataframe import transfer,shore_crane,forklift

    return forklift, transfer


@app.cell(hide_code=True)
def _(staged_electricity, transfer):
    _df = mo.sql(
        f"""
        WITH non_full_delivery AS (
            FROM transfer
            SELECT
                day_name,
                date,
                line,
                NULL AS vessel,  -- ✅ added to match full_delivery column count
                remarks AS customer,
                movement_type || ' of ' || status || ' ' || "type" || ' ' || size || ' by ' || driver AS operation_type,
                COUNT(*) AS number_of_service,
                SUM(shifting_price) AS shifting_price,
                SUM(haulage_price) AS haulage_price,
                CASE WHEN movement_type = 'Shifting' THEN 'INTERNAL SHIFTING' ELSE 'HAULAGE' END AS service,
                CASE
                    WHEN status = 'Empty' AND remarks IN ('MAERSKLINE', 'CMA CGM', 'IOT') THEN 'MONTHLY'
                    WHEN status = 'Full' AND remarks = 'MAERSKLINE' THEN 'OSS'
                    WHEN remarks = 'IOT' THEN 'MONTHLY'
                    WHEN movement_type = 'Delivery' AND status = 'Full' AND "type" = 'Reefer' THEN 'STO'
                    WHEN movement_type = 'Collection' AND status = 'Full' AND "type" = 'Reefer' THEN 'SI'
                    WHEN "type" = 'Dry' THEN 'SI'
                    ELSE 'SI'
                END AS sub_type,
                sub_type AS report_type
            WHERE
                remarks NOT IN ('IPHS')
                AND remarks <> 'CCCS'
                AND movement_type <> 'Shifting'
                AND NOT (movement_type = 'Delivery' AND status = 'Full')  -- ✅ fixed logic
            GROUP BY ALL
            ORDER BY date
        ),

        full_delivery AS (
            WITH
            elec AS (FROM staged_electricity),
            trans AS (
                FROM transfer
                WHERE movement_type = 'Delivery' AND status = 'Full'
            )
            SELECT
                t.day_name,
                t.date,
                t.line,
                e.vessel_client AS vessel,
                t.remarks AS customer,
                t.movement_type || ' of ' || t.status || ' ' || t."type" || ' ' || t.size || ' by ' || t.driver AS operation_type,
                COUNT(*) AS number_of_service,
                SUM(t.shifting_price) AS shifting_price,
                SUM(t.haulage_price) AS haulage_price,
                CASE WHEN movement_type = 'Shifting' THEN 'INTERNAL SHIFTING' ELSE 'HAULAGE' END AS service,
                CASE
                    WHEN status = 'Full' AND remarks = 'MAERSKLINE' AND e.operation_type = 'CCCS Stuffing - Basic OSS' THEN 'CCCS OSS'
                    WHEN status = 'Full' AND remarks = 'MAERSKLINE' AND e.operation_type = 'Basic OSS' THEN 'BASIC OSS'
                    WHEN status = 'Full' AND remarks = 'MAERSKLINE' AND e.operation_type = 'Full OSS' THEN 'FULL OSS'
                    WHEN remarks = 'IOT' THEN 'MONTHLY'
                    WHEN movement_type = 'Delivery' AND status = 'Full' AND "type" = 'Reefer' THEN 'STO'
                    WHEN movement_type = 'Collection' AND status = 'Full' AND "type" = 'Reefer' THEN 'SI'
                    WHEN "type" = 'Dry' THEN 'SI'
                    ELSE 'SI'
                END AS sub_type,
                CASE WHEN sub_type LIKE '%OSS%' THEN 'OSS' ELSE sub_type END AS report_type
            FROM trans t
            LEFT JOIN elec e ON e.period_end = t.date AND e.container_number = t.container_number
            WHERE
                remarks NOT IN ('IPHS')
                AND remarks <> 'CCCS'
                AND movement_type <> 'Shifting'
            GROUP BY ALL
            ORDER BY date
        )

        FROM non_full_delivery
        UNION ALL
        FROM full_delivery
        ORDER BY date
        """
    )
    return


@app.cell(hide_code=True)
def _():
    _df = mo.sql(
        f"""
        FROM
            shore_crane
        SELECT
            day_name,
            date,
            customer AS vessel,
            invoiced_to AS customer,
            operation_type,
            SUM("hours") AS billable_hours,
            SUM(overtime_hours) AS overtime_hours,
            unit_price,
            SUM(total_price) AS total_price,
            'SHORE CRANE RENTAL' AS service,
            CASE WHEN customer = 'MAERSKLINE' AND 
        GROUP BY ALL
        ORDER BY
            date
        """
    )
    return


@app.cell(hide_code=True)
def _(forklift):
    _df = mo.sql(
        f"""
        FROM (
            FROM forklift
            SELECT
                "day" AS day_name,
                date,
                customer AS vessel,
                invoiced_in AS customer,
                CEIL(SUM(normal_hours) / 60.0) * 30 AS normal_hours,
                CEIL(SUM(overtime_150) / 60.0) * 45 AS overtime_150,
                CEIL(SUM(overtime_200) / 60.0) * 60 AS overtime_200,
                'FORKLIFT RENTAL' AS service,
                CASE
                    WHEN invoiced_in IN ('HARTSWATER LIMITED','ECHEBASTAR','SAPMER SA','SAPMER','CCCS')
                    THEN 'MONTHLY'
                    ELSE 'STO'
                END AS sub_type,
                sub_type AS report_type
            GROUP BY ALL
        )
        SELECT
            *,
            (normal_hours + overtime_150 + overtime_200) AS total_price
        ORDER BY date
        """
    )
    return


@app.cell
def _():
    operations_url = "https://docs.google.com/spreadsheets/d/1PvTkl6DYZdhtaiNshz0qwtSPxC8S1OOeu905NmhFKNs/edit?gid=1431624134#gid=1431624134"

    master_url = "https://docs.google.com/spreadsheets/d/1ai-zQMtbPUx0LeQeLmXcpPgKvL5cyvwDfSJqRzxfUQg/edit?gid=0#gid=0"

    berth_url = "https://docs.google.com/spreadsheets/d/1xNh4SiP_xw8Ck1baLhFazBsmMLIbOHjF4tg_M89efvI/edit?gid=303791708#gid=303791708"

    well_tab = "WelltoWell"
    additional_tab = "AdditionalStevedores"
    extra_men_tab = "ExtraMen"
    tare_tab = "RawData"

    customer_tab = "Client"
    return (
        additional_tab,
        customer_tab,
        extra_men_tab,
        master_url,
        operations_url,
        tare_tab,
        well_tab,
    )


@app.cell
def _(customer_tab, master_url, scan_google_sheet):
    customer_df = scan_google_sheet(url=master_url,sheet_name=customer_tab)
    return (customer_df,)


@app.cell
def _(operations_url, scan_google_sheet, well_tab):
    well_df = scan_google_sheet(url=operations_url,sheet_name=well_tab)
    return (well_df,)


@app.cell
def _(customer_df, well_df):
    _df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    well_df
                SELECT
                    "Day" AS day_name,
                    "Date" AS date,
                    "Vessel" AS vessel,
                    "Tonnage" AS tonnage
            WHERE YEAR(date) = 2026
            ),
            client AS (
                FROM
                    customer_df
                SELECT
                    "Vessel/Client" AS vessel,
                    "Customer" AS customer
                WHERE
                    "Type" = 'PURSEINER'
            )
        SELECT
            d.*,
            c.customer,
            CASE WHEN day_name IN ('Sun','PH') THEN 17 * tonnage * 1.5 ELSE 17 * tonnage END AS total_price,
            'WELL TO WELL' AS service,
            'STO' AS sub_type,
            sub_type AS report_type
        FROM
            data_ d
            LEFT JOIN client c ON c.vessel = d.vessel
        """
    )
    return


@app.cell
def _(additional_tab, operations_url, scan_google_sheet):
    additional_df = scan_google_sheet(url=operations_url,sheet_name=additional_tab)
    return (additional_df,)


@app.cell(hide_code=True)
def _(additional_df):
    _df = mo.sql(
        f"""
        FROM additional_df
        """
    )
    return


@app.cell(hide_code=True)
def _(additional_df, customer_df):
    _df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    additional_df
                SELECT
                    day_name,
                    date,
                    vessel,
                    tonnage,
                    "hours",
                    number_of_stevedores,
                    "hours" * CASE
                        WHEN number_of_stevedores = 'To check' THEN 0
                        ELSE COALESCE(CAST(number_of_stevedores AS INT), 0)
                    END * 8 AS total_price
                WHERE
                    total_price > 0
            ),
            client AS (
                FROM
                    customer_df
                SELECT
                    "Vessel/Client" AS vessel,
                    "Customer" AS customer
                WHERE
                    "Type" = 'PURSEINER'
            )
            SELECT d.*,
            c.customer,
            'ADDITIONAL OVERTIME' AS service,
            'STO' AS sub_type,
            sub_type AS report_type
        FROM
            data_ d
            LEFT JOIN client c ON c.vessel = d.vessel
        """
    )
    return


@app.cell
def _(extra_men_tab, operations_url, scan_google_sheet):
    extra_men_df = scan_google_sheet(url=operations_url,sheet_name=extra_men_tab)
    return (extra_men_df,)


@app.cell(hide_code=True)
def _(customer_df, extra_men_df):
    _df = mo.sql(
        f"""
        WITH
            data_ AS (
                FROM
                    extra_men_df
                SELECT
                    "Day" AS day_name,
                    "Date" AS date,
                    "Vessel" AS vessel,
                    "Tonnage" AS tonnage,
                    "ExtraMen" AS extra_men
                WHERE
                    extra_men > 0
            ),
            client AS (
                FROM
                    customer_df
                SELECT
                    "Vessel/Client" AS vessel,
                    "Customer" AS customer
                WHERE
                    "Type" = 'PURSEINER'
            )
        SELECT
            d.*,
            c.customer,
            CASE
                WHEN day_name IN ('Sun', 'PH') THEN CAST(d.tonnage * d.extra_men * 0.8 * 1.5 AS DECIMAL)
                ELSE CAST(d.tonnage * d.extra_men * 0.8 AS DECIMAL)
            END AS total_price,
            'EXTRA MEN' AS service,
            'STO' AS sub_type,
            sub_type AS report_type
        FROM
            data_ d
            LEFT JOIN client c ON c.vessel = d.vessel
        """
    )
    return


@app.cell
def _(operations_url, scan_google_sheet, tare_tab):
    tare_df = scan_google_sheet(url=operations_url,sheet_name=tare_tab)
    return (tare_df,)


@app.cell(hide_code=True)
def _(customer_df, tare_df):
    _df = mo.sql(
        f"""
        WITH data_ AS (FROM
            tare_df
        SELECT
            "Day" AS day_name,
            "Date" AS date,
            UPPER(Vessel) AS vessel,
            1 AS rental,
            COUNT(DISTINCT "Side Working") AS number_of_side_working,
            STRING_AGG(DISTINCT "Side Working", ', ') AS side_working
        WHERE YEAR(date) = 2026
        GROUP BY ALL),
            client AS (
                FROM
                    customer_df
                SELECT
                    "Vessel/Client" AS vessel,
                    "Customer" AS customer
                WHERE
                    "Type" = 'PURSEINER'
            )
        SELECT
            d.*,
            c.customer,
            50 AS rental_price,
            number_of_side_working * 50 AS calibration_price,
            (rental_price + calibration_price) AS total_price,
            'TARE RENTAL AND CALIBRATION' AS service,
            'STO' AS sub_type,
            sub_type AS report_type
        FROM
            data_ d
            LEFT JOIN client c ON c.vessel = d.vessel

        """
    )
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
