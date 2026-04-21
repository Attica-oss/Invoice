import marimo

__generated_with = "0.23.1"
app = marimo.App(width="columns", app_title="Process Invoice")

with app.setup:
    from datetime import date, timedelta
    from enum import StrEnum

    import marimo as mo
    import polars as pl
    from scan_google_sheet import scan_google_sheet


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    # Process Invoices
    ---------

    * Load data from various Google Sheets into Polars LazyFrame
      * Use [DuckDB](https://duckdb.org/) to process these by adding day name, price, and customers.
      * Clean the data; flag invalid data and produce these into a dataframe view.
      * Currently, data is not getting persisted. However, we aim to use Parquet or DuckDB files.
      * There are some validations occurring in the Google Sheets; however, this is limited.
      * We are using strEnum to create enums to cast most columns.
    * Filter the datasets by date range and customer (vessels).
      * We have a Google Sheet that stores this information.
      * We aim to turn it into a `ReportMeta`
    * Save the data into an Excel Workbook using [Polars](https://docs.pola.rs/api/python/stable/reference/api/polars.DataFrame.write_excel.html) and [Xlsxwriter](https://xlsxwriter.readthedocs.io/working_with_polars.html)
    * Forma
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Casting Types
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Container Number Validator
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    #### STDU container.
    Only one in service currently.
    """)
    return


@app.cell
def _():
    class STDUContainer(PolarsEnum):
        """Cast the two STDU containers"""

        STDU6536343 = "STDU6536343"
        STDU6536338 = "STDU6536338"

    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    #### Container Validation Text Box Widget 🇬🇸
    """)
    return


@app.cell
def _():
    import re

    import anywidget
    import traitlets


    class ContainerValidationWidget(anywidget.AnyWidget):
        _esm = """
        function render({ model, el }) {
          const container = document.createElement("div");
          container.style.cssText = `
            font-family: monospace;
            padding: 14px 16px;
            border: 2px solid #555;
            border-radius: 6px;
            display: inline-block;
            transition: border-color 0.2s;
          `;

          const label = document.createElement("label");
          label.textContent = "Container Number";
          label.style.cssText = "display: block; font-weight: bold; margin-bottom: 6px; font-size: 13px;";

          const input = document.createElement("input");
          input.type = "text";
          input.placeholder = "e.g. MSCU1234565";
          input.style.cssText = `
            width: 200px;
            padding: 7px 10px;
            font-size: 14px;
            text-transform: uppercase;
            border: none;
            outline: none;
            font-family: monospace;
            letter-spacing: 1px;
            background: transparent;
            color: inherit;
          `;

          const validationMsg = document.createElement("div");
          validationMsg.style.cssText = "margin-top: 8px; font-size: 13px; font-weight: bold;";

          const details = document.createElement("div");
          details.style.cssText = `
            margin-top: 8px;
            font-size: 12px;
            opacity: 0.7;
            border-top: 1px solid #555;
            padding-top: 6px;
          `;

          function updateValidation() {
            const isValid = model.get("is_valid");
            const msg = model.get("validation_message");
            const owner = model.get("owner_code");
            const equip = model.get("equipment_category");
            const serial = model.get("serial_number");
            const check = model.get("check_digit");

            validationMsg.textContent = msg;

            if (!msg) {
              validationMsg.style.color = "";
              container.style.borderColor = "#555";
            } else if (isValid) {
              validationMsg.style.color = "#4caf50";
              container.style.borderColor = "#4caf50";
            } else {
              validationMsg.style.color = "#f44336";
              container.style.borderColor = "#f44336";
            }

            if (owner) {
              details.innerHTML = `
                <b>Owner:</b> ${owner} &nbsp;|&nbsp;
                <b>Category:</b> ${equip} &nbsp;|&nbsp;
                <b>Serial:</b> ${serial} &nbsp;|&nbsp;
                <b>Check Digit:</b> ${check}
              `;
            } else {
              details.innerHTML = "";
            }
          }

          input.addEventListener("input", (e) => {
            const value = e.target.value.toUpperCase();
            e.target.value = value;
            model.set("container_number", value);
            model.save_changes();
          });

          model.on("change:is_valid", updateValidation);
          model.on("change:validation_message", updateValidation);

          container.appendChild(label);
          container.appendChild(input);
          container.appendChild(validationMsg);
          container.appendChild(details);
          el.appendChild(container);

          updateValidation();
        }
        export default { render };
        """

        container_number = traitlets.Unicode("").tag(sync=True)
        is_valid = traitlets.Bool(False).tag(sync=True)
        validation_message = traitlets.Unicode("").tag(sync=True)
        owner_code = traitlets.Unicode("").tag(sync=True)
        equipment_category = traitlets.Unicode("").tag(sync=True)
        serial_number = traitlets.Unicode("").tag(sync=True)
        check_digit = traitlets.Unicode("").tag(sync=True)

        def __init__(self, **kwargs):
            super().__init__(**kwargs)
            self.observe(self._validate_container, names=["container_number"])

        def _validate_container(self, change):
            container = change["new"].strip().upper()

            if not container:
                self.is_valid = False
                self.validation_message = ""
                self.owner_code = ""
                self.equipment_category = ""
                self.serial_number = ""
                self.check_digit = ""
                return

            pattern = r"^([A-Z]{3})([UJZ])(\d{6})(\d)$"
            match = re.match(pattern, container)

            if not match:
                self.is_valid = False
                self.validation_message = (
                    "Invalid format. Expected: 3 owner letters + "
                    "equipment category (U/J/Z) + 6 digits + 1 check digit"
                )
                return

            owner, equipment, serial, check = match.groups()
            calculated_check = self._calculate_check_digit(
                owner + equipment + serial
            )

            self.owner_code = owner
            self.equipment_category = self._get_equipment_name(equipment)
            self.serial_number = serial
            self.check_digit = check

            if str(calculated_check) != check:
                self.is_valid = False
                self.validation_message = f"Invalid check digit. Expected {calculated_check}, got {check}"
            else:
                self.is_valid = True
                self.validation_message = "✓ Valid container number"

        def _calculate_check_digit(self, code: str) -> int:
            letter_values = {
                "A": 10,
                "B": 12,
                "C": 13,
                "D": 14,
                "E": 15,
                "F": 16,
                "G": 17,
                "H": 18,
                "I": 19,
                "J": 20,
                "K": 21,
                "L": 23,
                "M": 24,
                "N": 25,
                "O": 26,
                "P": 27,
                "Q": 28,
                "R": 29,
                "S": 30,
                "T": 31,
                "U": 32,
                "V": 34,
                "W": 35,
                "X": 36,
                "Y": 37,
                "Z": 38,
            }
            total = sum(
                (letter_values[c] if c.isalpha() else int(c)) * (2**i)
                for i, c in enumerate(code)
            )
            return (total % 11) % 10

        def _get_equipment_name(self, code: str) -> str:
            categories = {
                "U": "Freight Container",
                "J": "Detachable Freight Container Equipment",
                "Z": "Trailer and Chassis",
            }
            return f"{code} ({categories.get(code, 'Unknown')})"

    return (ContainerValidationWidget,)


@app.cell
def _(ContainerValidationWidget):
    myc = ContainerValidationWidget()
    myc
    return (myc,)


@app.cell
def _(myc):
    myc.is_valid
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Helper function
        * to clean transfer data, and keep only unique containers numbers as from 2025
    """)
    return


@app.cell
def _(TransferMovementType):
    def keep_only_collection_in_transfer(lf: pl.LazyFrame) -> pl.LazyFrame:
        return (
            lf.filter(
                pl.col("movement_type")
                .eq(TransferMovementType.COLLECTION)
                .and_(pl.col("date").dt.year().ge(2025))
            )
            .select(pl.col("container_number"), pl.col("line"))
            .unique()
        )

    return (keep_only_collection_in_transfer,)


@app.cell
def _(keep_only_collection_in_transfer):
    all_containers = scan_google_sheet(
        sheet_id="1O8K26c7CqLSdLr-f2gvZliDpBn9ArxXvj9tEJy-ElUg",
        sheet_name="Transfer",
    ).pipe(keep_only_collection_in_transfer)
    return (all_containers,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### IOT SOC
    """)
    return


@app.cell(hide_code=True)
def _(all_containers):
    iot_soc_list = mo.sql(
        f"""
        FROM all_containers
            SELECT container_number
        WHERE line = 'IOT'
        """,
        output=False,
    )
    return (iot_soc_list,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Other Reefers
    """)
    return


@app.cell(hide_code=True)
def _(all_containers):
    _df = mo.sql(
        f"""
        FROM all_containers
        WHERE line IN ('CMA CGM','MAERSK')
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ---
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Polars Enum

    - Great for type casting
    """)
    return


@app.class_definition
class PolarsEnum(StrEnum):
    """StrEnum base with Polars helpers."""

    @classmethod
    def list_all(cls) -> list[str]:
        """All allowed string values (in enum order)."""
        return [m.value for m in cls]

    @classmethod
    def enum_dtype(cls) -> pl.Enum:
        """Polars Enum dtype for casting/validation."""
        return pl.Enum(cls.list_all())

    @classmethod
    def has_value(cls, value: str) -> bool:
        """Fast membership check against allowed values."""
        return value in cls._value2member_map_

    @classmethod
    def parse(cls, value: str) -> PolarsEnum:
        """
        Parse an exact value into the enum.
        Raises ValueError if invalid.
        """
        try:
            return cls(value)
        except Exception as e:
            raise ValueError(
                f"Invalid {cls.__name__}: {value!r}. Allowed: {cls.list_all()}"
            ) from e

    @classmethod
    def normalize(cls, value: str) -> str:
        """
        Normalize user input to the canonical enum value (case/whitespace tolerant).
        Returns the canonical string value (not the enum member).
        """
        if value is None:
            raise ValueError(f"{cls.__name__} cannot be None")

        v = value.strip()
        # exact match first
        if cls.has_value(v):
            return v

        # case-insensitive match
        lower_map = {m.value.lower(): m.value for m in cls}
        key = v.lower()
        if key in lower_map:
            return lower_map[key]

        raise ValueError(
            f"Invalid {cls.__name__}: {value!r}. Allowed: {cls.list_all()}"
        )

    @classmethod
    def lit(cls, value: str) -> pl.Expr:
        """Polars literal cast to this enum dtype."""
        return pl.lit(cls.normalize(value)).cast(cls.enum_dtype())

    @classmethod
    def cast_col(cls, col: str | pl.Expr) -> pl.Expr:
        """Cast a column/expression to this enum dtype."""
        expr = pl.col(col) if isinstance(col, str) else col
        return expr.cast(cls.enum_dtype())


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Service Type Enum

    > To aid in fitering services that are grouped under a common name
    """)
    return


@app.class_definition
class ServiceGroupType(PolarsEnum):
    """Kind of Services by Group"""

    PTI = "PTI"
    NET_LIST = "Net List"
    CROSS_STUFFING = "Cross Stuffing"
    SALT = "Salt"
    STEVEDORING = "Stevedoring"
    BY_CATCH = "By Catch"
    DEPOT = "Depot"
    BIN_DISPATCH = "Bin Dispatch"
    CCCS = "CCCS"
    COLD_STORE_STUFFING = "Cold Store Stuffing"
    ELECTRICITY = "Electricity"
    TRANSFER = "Transfer"
    BERTH_DUES = "Berth Dues"


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Movement Type Enum
    """)
    return


@app.cell
def _():
    class TransferMovementType(PolarsEnum):
        """Kind of transfers"""

        DELIVERY = "Delivery"
        COLLECTION = "Collection"
        SHIFTING = "Shifting"


    class ColdStoreMovementType(PolarsEnum):
        """Cold Store Movement NOTE INTERVAL is not valid"""

        IN = "IN"
        OUT = "OUT"
        INTERNAL = "INTERNAL"

    return (TransferMovementType,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Vessel Type Enum
    """)
    return


@app.class_definition
class VesselType(PolarsEnum):
    """Vessel Type"""

    PURSEINER = "PURSEINER"
    CARGO_VESSEL = "CARGO VESSEL"
    SUPPLY_VESSEL = "SUPPLY VESSEL"
    LONGLINER = "LONGLINER"
    MILITARY_VESSEL = "MILITARY VESSEL"


@app.cell(column=1, hide_code=True)
def _():
    mo.md(r"""
    # Configs

    Google Sheets Urls and tab name (sheet name)
    """)
    return


@app.cell
def _(dataclass):
    import tomllib
    from pathlib import Path
    from types import MappingProxyType
    from typing import Mapping


    DEFAULT_CONFIG_PATH = Path(__file__).parent / "sheets.toml"


    def _template(key: str) -> str:
        return f"""\
    # Sheet configurations. Add one [section] per sheet group.
    #
    # [{key}]
    # url = "https://docs.google.com/spreadsheets/d/YOUR_SHEET_ID/edit"
    # sheets = {{ price = "Price", client = "Client" }}
    """


    @dataclass(frozen=True, slots=True)
    class SheetConfig:
        url: str
        sheets: Mapping[str, str]

        def __repr__(self) -> str:
            return f"SheetConfig(sheets={list(self.sheets)})"

        @classmethod
        def from_toml(
            cls,
            key: str,
            path: str | Path = DEFAULT_CONFIG_PATH,
        ) -> "SheetConfig":
            path = Path(path)

            if not path.exists():
                path.write_text(_template(key))
                raise FileNotFoundError(
                    f"No config found. Created template at {path}. "
                    f"Uncomment and fill in the [{key}] section, then retry."
                )

            with open(path, "rb") as f:
                data = tomllib.load(f)

            if key not in data:
                raise KeyError(
                    f"Section [{key}] not found in {path}. "
                    f"Available sections: {list(data) or '(none)'}"
                )

            section = data[key]
            return cls(
                url=section["url"],
                sheets=MappingProxyType(dict(section["sheets"])),
            )

        def sheet(self, key: str) -> str:
            if key not in self.sheets:
                raise KeyError(
                    f"Sheet key {key!r} not in config. Available: {list(self.sheets)}"
                )
            return self.sheets[key]

    return Path, SheetConfig


@app.cell(column=2, hide_code=True)
def _():
    mo.md(r"""
    ## Price DataFrame
    """)
    return


@app.cell
def _(SheetConfig):
    master = SheetConfig.from_toml("master")
    return (master,)


@app.cell
def _(master):
    price_raw_dataf = scan_google_sheet(
        url=master.url, sheet_name=master.sheet("price")
    )
    return (price_raw_dataf,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Process the LazyFrame with SQL
    """)
    return


@app.cell
def _(price_raw_dataf):
    price_df = mo.sql(
        f"""
        FROM
            price_raw_dataf
        SELECT
            Service AS service,
            "Class" AS service_type,
            Price AS unit_price,
            "StartingDate" AS effective_date
        WHERE
            "EndingDate" = ''
        """,
        output=False,
    )
    return (price_df,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Helper function to filter price list
    """)
    return


@app.function
def filter_price_by_group(
    df: pl.DataFrame | pl.LazyFrame,
    service_group: ServiceGroupType,
):
    return df.filter(pl.col("service_type").eq(service_group))


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Example
    """)
    return


@app.cell
def _(price_df):
    price_df.pipe(filter_price_by_group, ServiceGroupType.TRANSFER)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### For only one service

    📓 To prepare a function for this
    """)
    return


@app.cell
def _(price_df):
    price_df.filter(pl.col("service").eq("Forklift"))
    return


@app.cell(column=3, hide_code=True)
def _():
    mo.md(r"""
    ## BC Service List
    """)
    return


@app.cell
def _():
    data_path = r"C:\Users\gmounac\Downloads\Items.xlsx"

    bc_service_list = pl.read_excel(data_path)
    return (bc_service_list,)


@app.cell(hide_code=True)
def _(bc_service_list):
    bc_list = mo.sql(
        f"""
        FROM bc_service_list
            SELECT "No.","Description","Sheet","Mapping"
        WHERE Report = 'STO'
        """
    )
    return (bc_list,)


@app.cell
def _():
    return


@app.cell(column=4, hide_code=True)
def _():
    mo.md(r"""
    # Customers
    """)
    return


@app.cell
def _(master):
    client_raw_dataf = scan_google_sheet(
        url=master.url, sheet_name=master.sheet("client")
    )
    return (client_raw_dataf,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    #### Vessel list ⛴️
    """)
    return


@app.cell
def _():
    select_vessel = mo.ui.text(label="Search Vessel: ")
    select_vessel
    return (select_vessel,)


@app.cell
def _():
    iotc_vessel = pl.read_excel(
        source="./rav_search_2026-04-17T16_05_23.033Z.xlsx"
    )
    return (iotc_vessel,)


@app.cell(hide_code=True)
def _(iotc_vessel, select_vessel):
    _df = mo.sql(
        f"""
        FROM iotc_vessel WHERE "Name" LIKE '%{select_vessel.value}%'
        """
    )
    return


@app.cell(hide_code=True)
def _(client_raw_dataf):
    vessel_df = mo.sql(
        f"""
        FROM
            client_raw_dataf
        SELECT
            "Vessel/Client" AS vessel,
            "Customer" AS ship_owner
        WHERE
            "Type" IN (
                '{VesselType.PURSEINER}',
                '{VesselType.CARGO_VESSEL}',
                '{VesselType.SUPPLY_VESSEL}',
                '{VesselType.LONGLINER}'
            )
        """,
        output=False,
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    #### Cargo Vessel
    """)
    return


@app.cell(hide_code=True)
def _(client_raw_dataf):
    cargo_vessel_df = mo.sql(
        f"""
        FROM client_raw_dataf
            SELECT "Vessel/Client" AS vessel,
                "Customer" AS ship_owner
        WHERE "Type" = '{VesselType.CARGO_VESSEL}'
        """
    )
    return (cargo_vessel_df,)


@app.cell
def _():
    return


@app.cell(column=5, hide_code=True)
def _():
    mo.md(r"""
    ## Public Holiday DataFrame
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    #### Config ⚗️
    """)
    return


@app.cell
def _():
    # from __future__ import annotations

    # from datetime import date
    from dataclasses import dataclass


    @dataclass(frozen=True, slots=True)
    class ContinuousHoliday:
        start_year: int
        month: int
        day: int
        name: str

        def applies(self, year: int) -> bool:
            return year >= self.start_year

        def to_date(self, year: int) -> date:
            return date(year, self.month, self.day)


    NEW_CONTINUOUS_HOLIDAYS = [
        ContinuousHoliday(
            start_year=2026, month=2, day=1, name="Abolition of Slavery"
        ),
    ]

    ONE_TIME_HOLIDAYS_BY_YEAR: dict[int, set[date]] = {
        2025: {
            date(2025, 10, 11),
            date(2025, 10, 13),
            date(2025, 10, 27),
        }
    }

    FIXED_HOLIDAYS = [
        (1, 1),
        (1, 2),
        (5, 1),
        (6, 18),
        (6, 29),
        (8, 15),
        (11, 1),
        (12, 8),
        (12, 25),
    ]
    return (
        FIXED_HOLIDAYS,
        NEW_CONTINUOUS_HOLIDAYS,
        ONE_TIME_HOLIDAYS_BY_YEAR,
        dataclass,
    )


@app.cell
def _(dataclass):
    # from __future__ import annotations

    # from dataclasses import dataclass
    # from datetime import date, timedelta
    # import polars as pl

    # from holiday_config import (
    #     FIXED_HOLIDAYS,
    #     NEW_CONTINUOUS_HOLIDAYS,
    #     ONE_TIME_HOLIDAYS_BY_YEAR,
    # )


    @dataclass(slots=True)
    class PublicHolidayCalendar:
        fixed_holidays: list[tuple[int, int]]
        continuous_holidays: list[dict]
        one_time_holidays_by_year: dict[int, set[date]]

        def get_fixed_holidays(self, year: int) -> set[date]:
            return {date(year, month, day) for month, day in self.fixed_holidays}

        def get_continuous_holidays(self, year: int) -> set[date]:
            holidays: set[date] = set()

            for holiday in self.continuous_holidays:
                if holiday.applies(year):
                    holidays.add(holiday.to_date(year))

            return holidays

        def get_one_time_holidays(self, year: int) -> set[date]:
            return self.one_time_holidays_by_year.get(year, set())

        def get_easter_related_holidays(self, year: int) -> set[date]:
            easter = self._calculate_easter_sunday(year)

            return {
                easter,  # Easter Sunday
                easter - timedelta(days=2),  # Good Friday
                easter - timedelta(days=1),  # Holy Saturday
                easter + timedelta(days=1),  # Easter Monday
                easter + timedelta(days=60),  # Corpus Christi
            }

        def get_religious_holidays(self, year: int) -> set[date]:
            return self.get_easter_related_holidays(year)

        def get_holidays(self, year: int) -> set[date]:
            holidays: set[date] = set()

            holidays |= self.get_fixed_holidays(year)
            holidays |= self.get_continuous_holidays(year)
            holidays |= self.get_one_time_holidays(year)
            holidays |= self.get_religious_holidays(year)

            holidays |= self._get_monday_after_sunday_holidays(holidays)

            return holidays

        def to_lazyframe(self, year: int) -> pl.LazyFrame:
            holidays = sorted(self.get_holidays(year))
            return pl.LazyFrame({"date": holidays}).with_columns(
                pl.lit("PH").alias("day_name")
            )

        @staticmethod
        def _get_monday_after_sunday_holidays(holidays: set[date]) -> set[date]:
            return {
                holiday + timedelta(days=1)
                for holiday in holidays
                if holiday.weekday() == 6
            }

        @staticmethod
        def _calculate_easter_sunday(year: int) -> date:
            a = year % 19
            b = year // 100
            c = year % 100
            d = (19 * a + b - b // 4 - ((b - (b + 8) // 25 + 1) // 3) + 15) % 30
            e = (32 + 2 * (b % 4) + 2 * (c // 4) - d - (c % 4)) % 7
            f = d + e - 7 * ((a + 11 * d + 22 * e) // 451) + 114
            month = f // 31
            day = f % 31 + 1
            return date(year, month, day)

    return (PublicHolidayCalendar,)


@app.cell
def _(
    FIXED_HOLIDAYS,
    NEW_CONTINUOUS_HOLIDAYS,
    ONE_TIME_HOLIDAYS_BY_YEAR: dict[int, set[date]],
    PublicHolidayCalendar,
):
    holiday_calendar = PublicHolidayCalendar(
        fixed_holidays=FIXED_HOLIDAYS,
        continuous_holidays=NEW_CONTINUOUS_HOLIDAYS,
        one_time_holidays_by_year=ONE_TIME_HOLIDAYS_BY_YEAR,
    )

    public_holidays_lf = holiday_calendar.to_lazyframe(2026)
    return (public_holidays_lf,)


@app.cell
def _():
    return


@app.cell
def _(public_holidays_lf):
    public_holiday_dates = public_holidays_lf
    return (public_holiday_dates,)


@app.cell(column=6, hide_code=True)
def _():
    mo.md(r"""
    # Salt Operation 🧂
    """)
    return


@app.cell
def _(SheetConfig):
    salt = SheetConfig.from_toml("salt")
    return (salt,)


@app.cell
def _(salt):
    salt_raw = scan_google_sheet(url=salt.url, sheet_name=salt.sheet("salt_ops"))
    return (salt_raw,)


@app.cell
def _(price_df):
    _df = mo.sql(
        f"""
        SELECT
                    service AS pricing_service,
                    unit_price
                FROM
                    price_df
                WHERE
                    service_type = '{ServiceGroupType.SALT}'
        """
    )
    return


@app.cell(hide_code=True)
def _(salt_raw):
    _df = mo.sql(
        f"""
        FROM salt_raw
        WHERE YEAR(date) = 2026 AND start_time <> ''
        """
    )
    return


@app.cell
def _():
    return


@app.cell(column=7, hide_code=True)
def _():
    mo.md(r"""
    # Shore Crane Rental Data
    """)
    return


@app.cell
def _():
    shore_crane_raw = scan_google_sheet(
        sheet_id="1O8K26c7CqLSdLr-f2gvZliDpBn9ArxXvj9tEJy-ElUg",
        sheet_name="ShoreCrane",
    )
    return (shore_crane_raw,)


@app.cell(hide_code=True)
def _(price_df, public_holiday_dates, shore_crane_raw):
    shore_crane_raw_df = mo.sql(
        f"""
        WITH unit_price AS (
            SELECT service, unit_price
            FROM price_df
            WHERE service = 'Shore Crane'
        ),

        ph_dates AS (
            SELECT
                CAST(date AS DATE) AS ph_date,
                day_name
            FROM public_holiday_dates
        ),

        raw AS (
            SELECT
                CAST("date" AS DATE) AS service_date,
               TRY_CAST(NULLIF(TRIM(CAST(start_time AS VARCHAR)), '') AS TIME)AS start_time,
               TRY_CAST(NULLIF(TRIM(CAST(end_time AS VARCHAR)), '') AS TIME) AS end_time,
                "hours",
                "overtime_hours",
                customer,
                location,
                operation_type,
                remarks,
                invoiced_to,
                'Shore Crane' AS service
            FROM shore_crane_raw
            WHERE YEAR("date") = 2026
        ),

        enriched AS (
            SELECT
                r.*,
                CASE
                    WHEN p.ph_date IS NOT NULL THEN p.day_name
                    ELSE STRFTIME(r.service_date, '%a')
                END AS day_name,
                CASE
                    WHEN remarks IS NULL OR remarks = '' THEN operation_type
                    ELSE operation_type || ' - ' || remarks
                END AS remarks_final
            FROM raw r
            LEFT JOIN ph_dates p
                ON r.service_date = p.ph_date
        ),

        validated AS (
            SELECT
                e.*,
                start_time IS NULL AS missing_start_time,
                end_time IS NULL AS missing_end_time,
                start_time IS NOT NULL
                    AND end_time IS NOT NULL
                    AND end_time < start_time AS end_before_start_time
            FROM enriched e
        ),

        durations AS (
            SELECT
                v.*,
                CASE
                    WHEN missing_start_time OR missing_end_time OR end_before_start_time THEN NULL
                    WHEN "hours" IS NULL THEN NULL
                    ELSE CEIL(EXTRACT(EPOCH FROM "hours") / 3600.0)
                END AS total_hours,
                CASE
                    WHEN missing_start_time OR missing_end_time OR end_before_start_time THEN NULL
                    WHEN "overtime_hours" IS NULL THEN NULL
                    ELSE CEIL(EXTRACT(EPOCH FROM "overtime_hours") / 3600.0)
                END AS overtime_hours_rounded
            FROM validated v
        ),

        normalized AS (
            SELECT
                d.*,
                CASE
                    WHEN total_hours IS NULL OR overtime_hours_rounded IS NULL THEN NULL
                    WHEN overtime_hours_rounded > total_hours THEN NULL
                    ELSE total_hours - overtime_hours_rounded
                END AS normal_hours,
                CASE
                    WHEN total_hours IS NOT NULL
                     AND overtime_hours_rounded IS NOT NULL
                     AND overtime_hours_rounded > total_hours
                    THEN TRUE
                    ELSE FALSE
                END AS overtime_exceeds_total
            FROM durations d
        ),

        priced AS (
            SELECT
                n.service_date,
                n.day_name,
                n.start_time,
                n.end_time,
                n.total_hours,
                n.overtime_hours_rounded AS overtime_hours,
                n.normal_hours,
                n.customer,
                n.location,
                n.remarks_final AS remarks,
                n.invoiced_to,
                u.unit_price AS base_unit_price,

                CASE
                    WHEN n.day_name IN ('Sun', 'PH')
                     AND n.service_date >= DATE '2026-03-01'
                        THEN ROUND(u.unit_price * 1.6, 3)
                    WHEN n.day_name IN ('Sun', 'PH')
                        THEN ROUND(u.unit_price * 1.5, 3)
                    ELSE ROUND(u.unit_price, 3)
                END AS unit_price,

                n.missing_start_time,
                n.missing_end_time,
                n.end_before_start_time,
                n.overtime_exceeds_total,

                CASE
                    WHEN n.missing_start_time THEN 'Missing start_time'
                    WHEN n.missing_end_time THEN 'Missing end_time'
                    WHEN n.end_before_start_time THEN 'end_time before start_time'
                    WHEN n.overtime_exceeds_total THEN 'overtime_hours exceeds total_hours'
                    ELSE 'OK'
                END AS validation_status
            FROM normalized n
            LEFT JOIN unit_price u
                ON u.service = n.service
        )

        FROM priced;
        """,
        output=False,
    )
    return (shore_crane_raw_df,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Invalid Record to Clean Up
    """)
    return


@app.cell(hide_code=True)
def _(shore_crane_raw_df):
    _df = mo.sql(
        f"""
        FROM shore_crane_raw_df
        WHERE validation_status <> 'OK'
        """
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    Valid Record to Process
    """)
    return


@app.cell(hide_code=True)
def _(shore_crane_raw_df):
    final_shore_crane_df = mo.sql(
        f"""
        SELECT
            day_name,
            service_date,
            start_time,
            end_time,
            total_hours,
            overtime_hours,
            customer,
            location,
            remarks,
            unit_price,
            CASE
                -- WHEN validation_status <> 'OK' THEN NULL

                WHEN day_name IN ('Sun', 'PH')
                 AND service_date >= DATE '2026-03-01'
                    THEN ROUND(
                        (base_unit_price * 1.6 * normal_hours) +
                        (base_unit_price * 2.1 * overtime_hours),
                        3
                    )

                WHEN day_name NOT IN ('Sun', 'PH')
                 AND service_date >= DATE '2026-03-01'
                    THEN ROUND(
                        (base_unit_price * 1.0 * normal_hours) +
                        (base_unit_price * 1.6 * overtime_hours),
                        3
                    )

                WHEN day_name IN ('Sun', 'PH')
                    THEN ROUND(
                        (base_unit_price * 1.5 * normal_hours) +
                        (base_unit_price * 2.0 * overtime_hours),
                        3
                    )

                ELSE ROUND(
                    (base_unit_price * 1.0 * normal_hours) +
                    (base_unit_price * 1.5 * overtime_hours),
                    3
                )
            END AS total_price,
            invoiced_to
        FROM shore_crane_raw_df
            WHERE validation_status = 'OK'
        ORDER BY service_date, start_time;
        """,
        output=False,
    )
    return (final_shore_crane_df,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Sample of the data (LIMIT 5;)
    """)
    return


@app.cell(hide_code=True)
def _(final_shore_crane_df):
    _df = mo.sql(
        f"""
        FROM final_shore_crane_df
        LIMIT 5
        """
    )
    return


@app.cell(column=8, hide_code=True)
def _():
    mo.md(r"""
    # Net List // Unloading Records
    """)
    return


@app.cell
def _(SheetConfig):
    operations = SheetConfig.from_toml("operations")
    miscellaneous = SheetConfig.from_toml("miscellaneous")
    stuffing = SheetConfig.from_toml("stuffing")
    return miscellaneous, operations, stuffing


@app.cell
def _(miscellaneous, operations, stuffing):
    net_list_raw = scan_google_sheet(
        url=operations.url, sheet_name=operations.sheet("net_list")
    )
    to_cold_store_via_truck_raw = scan_google_sheet(
        url=miscellaneous.url, sheet_name=miscellaneous.sheet("iphs_truck")
    )
    container_stuffing_raw = scan_google_sheet(
        url=stuffing.url, sheet_name=stuffing.sheet("ops_activity")
    )
    return container_stuffing_raw, net_list_raw, to_cold_store_via_truck_raw


@app.cell(hide_code=True)
def _(net_list_raw, to_cold_store_via_truck_raw):
    cold_store_adjusted_dataf = mo.sql(
        f"""
        WITH
            base AS (
                SELECT
                    "day" AS day_name,
                    date AS service_date,
                    'CCCS (' || REPLACE(REPLACE(customer, ' S.A.', ''), ' S.A', '') || ')' AS destination,
                    vessel,
                    CAST(total_tonnage - overtime_tonnage AS DECIMAL) AS tonnage,
                    overtime_tonnage,
                    "storage" AS storage_type
                FROM
                    to_cold_store_via_truck_raw
                WHERE
                    YEAR(date) = 2026
                    AND operation_type = 'To CCCS via Truck'
            ),
            normal AS (
                SELECT
                    * EXCLUDE (overtime_tonnage),
                    CASE
                        WHEN day_name IN ('Sun', 'PH') THEN 'overtime 150%'
                        ELSE 'normal hours'
                    END AS overtime
                FROM
                    base
            ),
            overtime AS (
                SELECT
                   day_name,
            service_date,
            destination,
            vessel,
            overtime_tonnage AS tonnage,
            storage_type,
                    CASE
                        WHEN day_name IN ('Sun', 'PH') THEN 'overtime 200%'
                        ELSE 'overtime 150%'
                    END AS overtime,

                FROM
                    base
                WHERE
                    overtime_tonnage > 0
            ),
            adjusted_cold_store AS (
                SELECT
                    *
                FROM
                    normal
                UNION ALL
                SELECT
                    *
                FROM
                    overtime
                ORDER BY
                    service_date,
                    vessel
            ),
            raw_cold_store AS (
                FROM
                    net_list_raw
                SELECT
                    "Date" AS service_date,
                    UPPER("Vessel") AS vessel,
                    "startTime" AS start_time,
                    "Container (Destination)" AS destination,
                    overtime,
                    "Storage" AS storage_type,
                    "endTime" AS end_time,
                    "Total Tonnage" AS total_tonnage
                WHERE
                    YEAR(date) = 2026
                    AND destination LIKE '%CCCS%'
            )

            SELECT r.*,
            a.tonnage
        FROM
            raw_cold_store r
            LEFT JOIN adjusted_cold_store a ON a.service_date = r.service_date
            AND r.vessel = a.vessel
            AND r.destination = a.destination
            AND r.storage_type = a.storage_type
            AND r.overtime = a.overtime
        """,
        output=False,
    )
    return (cold_store_adjusted_dataf,)


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Check the Cold Store Tonnage Against Genesis Data
    """)
    return


@app.cell(hide_code=True)
def _(cold_store_adjusted_dataf):
    cccs_tonnage_comparison = mo.sql(
        f"""
        FROM
            cold_store_adjusted_dataf
        SELECT
            service_date,
            vessel,
            destination,
            total_tonnage,
            tonnage,
           ROUND((tonnage - total_tonnage),3) AS var,
            CASE WHEN ROUND((tonnage - total_tonnage),3) > 0 THEN '🔺'
           WHEN ROUND((tonnage - total_tonnage),3) < 0 THEN '🔻'
            ELSE '🟰' END AS var_symbol
        ORDER BY var
        LIMIT 5
        """
    )
    return


@app.cell(hide_code=True)
def _(
    cargo_vessel_df,
    cold_store_adjusted_dataf,
    container_stuffing_raw,
    iot_soc_list,
    net_list_raw,
    price_df,
    public_holiday_dates,
):
    final_net_list_df = mo.sql(
        f"""
        WITH
            unit_price AS (
                SELECT
                    service AS pricing_service,
                    unit_price
                FROM
                    price_df
                WHERE
                    service_type = '{ServiceGroupType.NET_LIST}'
            ),
            ph_dates AS (
                SELECT
                    CAST(date AS DATE) AS ph_date,
                    day_name
                FROM
                    public_holiday_dates
            ),
            base AS (
                FROM
                    net_list_raw
                SELECT
                    "Date" AS service_date,
                    UPPER("Vessel") AS vessel,
                    "startTime" AS start_time,
                    "Container (Destination)" AS destination,
                    overtime,
                    "Storage" AS storage_type,
                    "endTime" AS end_time,
                    "Total Tonnage" AS total_tonnage
                WHERE
                    YEAR(date) = 2026
                    AND destination NOT LIKE '%CCCS%'
            ),
            cold_store AS (
                FROM
                    cold_store_adjusted_dataf
                SELECT
                    service_date,
                    vessel,
                    start_time,
                    destination,
                    overtime,
                    storage_type,
                    end_time,
                    tonnage
            ),
            stuffing AS (
                FROM
                    container_stuffing_raw
                SELECT
                    vessel_client AS vessel,
                    date_plugged AS service_date,
                    container_number AS destination,
                    CASE
                        WHEN operation_type LIKE '%Full%' THEN 'Full OSS'
                        WHEN operation_type LIKE '%Basic%' THEN 'Basic OSS'
                        ELSE 'Container Stuffing'
                    END AS service,
                    tonnage
                WHERE
                    (
                        YEAR(date_out) = 2026
                        OR date_out IS NULL
                    )
                    AND (
                        operation_type LIKE '%Basic OSS%'
                        OR operation_type LIKE '%Full OSS%'
                        OR operation_type LIKE '%Stuffing and Plugin%'
                    )
                    AND (
                        operation_type NOT LIKE '%CCCS Stuffing%'
                        AND operation_type NOT LIKE '%Cross%'
                    )
                    AND shipping_line <> 'IOT'
            ),
            grouped_data AS (
                FROM
                    base
                UNION ALL
                FROM
                    cold_store
                ORDER BY
                    service_date,
                    vessel
            ),
            add_day_name AS (
                SELECT
                    CASE
                        WHEN p.ph_date IS NOT NULL THEN p.day_name
                        ELSE STRFTIME(g.service_date, '%a')
                    END AS day_name,
                    g.*
                FROM
                    grouped_data g
                    LEFT JOIN ph_dates p ON g.service_date = p.ph_date
            ),
            add_stuffing_service AS (
                SELECT
                    a.*,
                    s.service AS service,
                    -- s.tonnage AS stuffing_tonnage,
                    -- s.service_date AS stuffing_service_date,
                    CASE
                        WHEN a.vessel = s.vessel THEN NULL
                        ELSE s.vessel
                    END AS remarks
                FROM
                    add_day_name a
                    LEFT JOIN LATERAL (
                        SELECT
                            s.service,
                            s.tonnage,
                            s.service_date,
                            s.vessel
                        FROM
                            stuffing s
                        WHERE
                            s.destination = a.destination
                            AND s.service_date >= a.service_date
                            AND s.service_date <= a.service_date + INTERVAL 1 DAY
                            AND (
                                s.vessel = a.vessel
                                OR (
                                    s.service = 'Basic OSS'
                                    AND s.vessel IN ('OCEAN BASKET', 'AMIRANTE')
                                )
                            )
                        ORDER BY
                            CASE
                                WHEN s.service_date = a.service_date THEN 0
                                ELSE 1
                            END,
                            CASE
                                WHEN s.vessel = a.vessel THEN 0
                                ELSE 1
                            END
                        LIMIT
                            1
                    ) s ON TRUE
            ),
            add_all_service AS (
                FROM
                    add_stuffing_service
                SELECT
                    day_name,
                    service_date,
                    vessel,
                    start_time,
                    destination,
                    overtime,
                    storage_type,
                    end_time,
                    total_tonnage,
                    CASE
                        WHEN service IS NOT NULL THEN service
                        WHEN destination LIKE '%IOT%'
                        OR destination = 'Unload to Quay'
                        OR destination IN (
                            FROM
                                iot_soc_list
                        ) THEN 'Unload to Quay'
                        WHEN UPPER(destination) IN (
                            FROM
                                cargo_vessel_df
                            SELECT
                                vessel
                        ) THEN 'Transhipment'
                        WHEN destination LIKE '%CCCS%' THEN 'Unload to CCCS'
                    END AS service,
                    remarks
                ORDER BY
                    service,
                    vessel
            ),
            add_pricing_service AS (
                FROM
                    add_all_service
                SELECT
                    day_name,
                    service_date,
                    vessel,
                    start_time,
                    destination,
                    overtime,
                    storage_type,
                    end_time,
                    total_tonnage,
                    service,
                    service || ' - ' || storage_type AS pricing_service,
                    remarks
            )
        SELECT
            s.* EXCLUDE (s.pricing_service, s.remarks),
            u.unit_price,
            CASE
                WHEN overtime = 'overtime 200%' THEN CAST(2.0 * u.unit_price * s.total_tonnage AS DECIMAL)
                WHEN overtime = 'overtime 150%' THEN CAST(1.5 * u.unit_price * s.total_tonnage AS DECIMAL)
                ELSE CAST(1.0 * u.unit_price * s.total_tonnage AS DECIMAL)
            END AS total_price,
            s.remarks
        FROM
            add_pricing_service s
            LEFT JOIN unit_price u ON u.pricing_service = s.pricing_service
        ORDER BY s.service_date,s.vessel
        """
    )
    return (final_net_list_df,)


@app.cell(hide_code=True)
def _(final_net_list_df):
    _df = mo.sql(
        f"""
        FROM final_net_list_df
        LIMIT 5
        """
    )
    return


@app.cell
def _(end_date, final_net_list_df, start_date, vessel_name):
    final_net_df = final_net_list_df.filter(
        (pl.col("service_date") >= start_date)
        & (pl.col("service_date") <= end_date)
        & (pl.col("vessel") == vessel_name)
    )
    return (final_net_df,)


@app.cell(hide_code=True)
def _(bc_list):
    _df = mo.sql(
        f"""
        FROM bc_list
        WHERE sheet = 'NetList'
        """
    )
    return


@app.cell(hide_code=True)
def _(bc_list, final_net_df):
    _df = mo.sql(
        f"""
        WITH
            final AS (
                FROM
                    final_net_df
            ),
            maped AS (
                FROM
                    bc_list
                WHERE
                    sheet = 'NetList'
            ),
            joined AS (
                SELECT
                    m."No.",
                    f.*,
                    CASE
                        WHEN f.overtime = 'normal hours' THEN 'STD'
                        WHEN f.overtime = 'overtime 150%' THEN '150%'
                        ELSE '200%'
                    END AS "Variant Code"
                FROM
                    final f
                    LEFT JOIN maped m ON m.Mapping = f.service || ' - ' || f.storage_type
            )

        FROM joined
        SELECT 'Item' AS "Type","No.",' ' AS Description,"Variant Code",CAST(SUM(total_tonnage) AS DECIMAL) AS Quantity
        GROUP BY ALL
        """
    )
    return


@app.cell
def _(Path):
    t_cols = {"total_tonnage": "sum", "total_price": "sum"}


    def save_to_excel(
        df: pl.DataFrame, file_path: str | Path, sheet: str, totals: dict[str, str]
    ) -> None:
        """Save raw to Excel"""
        return df.write_excel(
            workbook=file_path,
            worksheet=sheet,
            table_name=str(sheet).strip().lower().replace(" ", "_"),
            table_style="Table Style Medium 2",
            dtype_formats={pl.Date: "yyyy-mm-dd", pl.Time: "hh:mm"},
            column_totals=totals,
        )

    return


@app.cell
def _(final_net_df):
    final_net_df
    return


@app.cell
def _(final_net_df):
    final_net_df.write_excel(
        "shore_crane_report.xlsx",
        worksheet="Net List",
        table_style="Table Style Medium 2",
        dtype_formats={
            pl.Date: "yyyy-mm-dd",
            pl.Time: "hh:mm",
        },
        column_totals={"total_tonnage": "sum", "total_price": "sum"},
    )
    return


@app.cell
def _(format_generic_report, meta):
    format_generic_report(
        xlsx_path="shore_crane_report.xlsx",
        sheet_name="Net List",
        report_meta=meta,
        footer_text="Net List",
        currency_cols=["unit_price", "total_price"],
        date_cols=["service_date"],
        time_cols=["start_time", "end_time"],
        right_headers=["unit_price", "total_price", "total_tonnage"],
        center_headers=["day_name", "service_date", "start_time", "end_time"],
        header_color="595959",
        number_format_map={
            "unit_price": "#,##0.00",
            "total_tonnage": "#,###0.000",
        },
        width_overrides={
            "day_name": 10,
            "service_date": 14,
            "vessel": 18,
            "start_time": 11,
            "destination": 20,
            "overtime": 16,
            "storage_type": 14,
            "end_time": 11,
            "total_tonnage": 14,
            "service": 18,
            "unit_price": 12,
            "total_price": 14,
            "remarks": 18,
        },
    )
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ## Excel Writer
    """)
    return


@app.cell
def _():
    import xlsxwriter

    return (xlsxwriter,)


@app.cell
def _():
    return


@app.cell
def _(
    end_date,
    final_net_df,
    final_shore_crane_df,
    start_date,
    vessel_name,
    xlsxwriter,
):
    xlsx_path = "combined_report.xlsx"

    workbook = xlsxwriter.Workbook(xlsx_path)

    final_net_df.write_excel(
        workbook=workbook,
        worksheet="Net List",
        table_style="Table Style Medium 2",
        dtype_formats={
            pl.Date: "yyyy-mm-dd",
            pl.Time: "hh:mm",
        },
        column_totals={"total_tonnage": "sum", "total_price": "sum"},
    )

    shore_crane_filtered = (
        final_shore_crane_df.filter(
            (pl.col("service_date") >= start_date)
            & (pl.col("service_date") <= end_date)
            & (pl.col("customer") == vessel_name)
            & (pl.col("invoiced_to") != "MAERSKLINE")
        )
        .rename(
            {
                "day_name": "Day Name",
                "service_date": "Date",
                "start_time": "Start Time",
                "end_time": "End Time",
                "total_hours": "Hours",
                "overtime_hours": "Overtime Hours",
                "customer": "Client",
                "unit_price": "Price / Hour ($)",
                "remarks": "Operation",
                "total_price": "Total Price ($)",
            }
        )
        .drop(["invoiced_to"])
        .select(
            [
                "Day Name",
                "Date",
                "Start Time",
                "End Time",
                "Hours",
                "Overtime Hours",
                "Client",
                "Price / Hour ($)",
                "Operation",
                "Total Price ($)",
            ]
        )
    )

    shore_crane_filtered.write_excel(
        workbook=workbook,
        worksheet="Shore Crane",
        table_style="Table Style Medium 2",
        dtype_formats={
            pl.Date: "yyyy-mm-dd",
            pl.Time: "hh:mm",
        },
        column_formats={
            "Price / Hour ($)": "$#,##0.000",
            "Total Price ($)": "$#,##0.000",
            "Hours": "0",
            "Overtime Hours": "0",
        },
        column_totals={"Hours": "sum", "Total Price ($)": "sum"},
    )

    workbook.close()
    return (xlsx_path,)


@app.cell
def _(format_generic_report, meta, xlsx_path):
    format_generic_report(
        xlsx_path=xlsx_path,
        sheet_name="Net List",
        report_meta=meta,
        footer_text="Net List",
        currency_cols=["unit_price", "total_price"],
        date_cols=["service_date"],
        time_cols=["start_time", "end_time"],
        right_headers=["unit_price", "total_price", "total_tonnage"],
        center_headers=["day_name", "service_date", "start_time", "end_time"],
        header_color="595959",
        number_format_map={
            "unit_price": "#,##0.00",
            "total_tonnage": "#,###0.000",
        },
        width_overrides={
            "day_name": 10,
            "service_date": 14,
            "vessel": 18,
            "start_time": 11,
            "destination": 20,
            "overtime": 16,
            "storage_type": 14,
            "end_time": 11,
            "total_tonnage": 14,
            "service": 18,
            "unit_price": 12,
            "total_price": 14,
            "remarks": 18,
        },
    )


    format_generic_report(
        xlsx_path=xlsx_path,
        sheet_name="Shore Crane",
        report_meta=meta,
        footer_text="Shore Crane",
        currency_cols=["Price / Hour ($)", "Total Price ($)"],
        date_cols=["Date"],
        time_cols=["Start Time", "End Time"],
        int_cols=["Hours", "Overtime Hours"],
        right_headers=["Price / Hour ($)", "Total Price ($)", "Hours"],
        center_headers=["Day Name", "Date", "Start Time", "End Time"],
        header_color="1F4E78",
        number_format_map={
            "Price / Hour ($)": "#,##0.00",
            "Total Price ($)": "#,###0.000",
            "Hours": '_-* #,##0_-;-* #,##0_-;_-* " - "??_-;_-@_-',
            "Overtime Hours": '_-* #,##0_-;-* #,##0_-;_-* " - "??_-;_-@_-',
        },
        width_overrides={
            "Day Name": 11,
            "Date": 13,
            "Start Time": 12,
            "End Time": 12,
            "Hours": 8,
            "Overtime Hours": 15,
            "Client": 22,
            "Price / Hour ($)": 16,
            "Operation": 28,
            "Total Price ($)": 16,
        },
    )
    return


@app.cell(column=9, hide_code=True)
def _():
    mo.md(r"""
    ### Summary Sheet
    """)
    return


@app.cell
def _(xlsxwriter):
    summary_data = {
        "start_date": "2026-01-01",
        "end_date": "2026-01-08",
        "vessel": "Izaro",
        "owner": "HARTSWATER",
        "sto_number": "STO-2500221",
        "collection_delivery": "-",
        "move_1": "- Move",
        "collection_delivery_2": "-",
        "move_2": "- Move",
        "total_moves": "-",
        "total_tonnage": "ERROR",
        "total_container_stuffed": 7,
        "forklift_services": 1,
        "salt_operations": "-",
        "pallets": "-",
    }


    def write_summary_sheet(xlsx_path, summary_data: dict) -> None:
        workbook = xlsxwriter.Workbook(xlsx_path)
        ws = workbook.add_worksheet("Summary")

        ws.hide_gridlines(2)

        # -----------------------------
        # Column widths / row heights
        # -----------------------------
        ws.set_column("A:A", 3)
        ws.set_column("B:B", 24)
        ws.set_column("C:C", 14)
        ws.set_column("D:D", 14)
        ws.set_column("E:E", 8)
        ws.set_column("F:F", 14)
        ws.set_column("G:G", 4)

        for r in range(0, 16):
            ws.set_row(r, 22)

        # -----------------------------
        # Formats
        # -----------------------------
        outer_border = workbook.add_format(
            {
                "border": 1,
                "bg_color": "#F2F2F2",
            }
        )

        label_fmt = workbook.add_format(
            {
                "bold": True,
                "align": "right",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
            }
        )

        center_fmt = workbook.add_format(
            {
                "align": "center",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
            }
        )

        green_title_fmt = workbook.add_format(
            {
                "bg_color": "#C6EFCE",
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_size": 15,
            }
        )

        green_value_fmt = workbook.add_format(
            {
                "bg_color": "#C6EFCE",
                "align": "center",
                "valign": "vcenter",
                "font_size": 12,
            }
        )

        metric_label_fmt = workbook.add_format(
            {
                "bold": True,
                "align": "right",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "font_size": 12,
            }
        )

        metric_value_fmt = workbook.add_format(
            {
                "align": "center",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "font_size": 12,
            }
        )

        metric_red_fmt = workbook.add_format(
            {
                "font_color": "#FF0000",
                "align": "center",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "font_size": 12,
            }
        )

        unit_fmt = workbook.add_format(
            {
                "italic": True,
                "align": "left",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "font_size": 12,
            }
        )

        underline_center_fmt = workbook.add_format(
            {
                "align": "center",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "bottom": 1,
            }
        )

        underline_right_fmt = workbook.add_format(
            {
                "align": "right",
                "valign": "vcenter",
                "bg_color": "#F2F2F2",
                "bottom": 1,
            }
        )

        # -----------------------------
        # Paint background block
        # -----------------------------
        for row in range(0, 15):
            for col in range(0, 7):
                ws.write_blank(row, col, None, outer_border)

        # -----------------------------
        # Outer box only
        # -----------------------------
        ws.conditional_format(
            "A1:G15", {"type": "no_blanks", "format": outer_border}
        )
        ws.conditional_format("A1:G15", {"type": "blanks", "format": outer_border})

        # -----------------------------
        # Header section
        # -----------------------------
        ws.write("B1", "Date:", label_fmt)
        ws.write("C1", summary_data.get("start_date", ""), center_fmt)
        ws.write("E1", "to", center_fmt)
        ws.write("F1", summary_data.get("end_date", ""), center_fmt)

        ws.write("B2", "Vessel:", label_fmt)
        ws.merge_range("C2:F2", summary_data.get("vessel", ""), green_value_fmt)

        ws.write("B3", "Ship Owner/Operator:", label_fmt)
        ws.merge_range("C3:F3", summary_data.get("owner", ""), green_title_fmt)

        ws.write("B4", "STO Number", label_fmt)
        ws.merge_range(
            "C4:F4", summary_data.get("sto_number", ""), green_value_fmt
        )

        # -----------------------------
        # Middle haulage section
        # -----------------------------
        ws.write("B6", "Haulage:", metric_label_fmt)

        ws.write(
            "C6",
            summary_data.get("collection_delivery", "-"),
            underline_center_fmt,
        )
        ws.write("F6", summary_data.get("move_1", "- Move"), underline_center_fmt)

        ws.write(
            "C7",
            summary_data.get("collection_delivery_2", "-"),
            underline_center_fmt,
        )
        ws.write("F7", summary_data.get("move_2", "- Move"), underline_center_fmt)

        ws.write("C8", "Total Moves", metric_label_fmt)
        ws.write("F8", summary_data.get("total_moves", "-"), underline_center_fmt)

        # -----------------------------
        # Lower metrics
        # -----------------------------
        ws.write("B10", "Total Tonnage:", metric_label_fmt)
        ws.write("C10", summary_data.get("total_tonnage", "ERROR"), metric_red_fmt)
        ws.write("D10", "Tons", unit_fmt)

        ws.write("B11", "Total Container Stuffed:", metric_label_fmt)
        ws.write(
            "C11", summary_data.get("total_container_stuffed", 0), metric_value_fmt
        )
        ws.write("D11", "Containers", unit_fmt)

        ws.write("B12", "Forklift Services:", metric_label_fmt)
        ws.write("C12", summary_data.get("forklift_services", 0), metric_value_fmt)
        ws.write("D12", "Hour", unit_fmt)

        ws.write("B13", "Salt Operations:", metric_label_fmt)
        ws.write("C13", summary_data.get("salt_operations", "-"), metric_value_fmt)
        ws.write("D13", "Ton", unit_fmt)

        ws.write("B14", "Pallets:", metric_label_fmt)
        ws.write("C14", summary_data.get("pallets", "-"), metric_value_fmt)
        ws.write("D14", "Pallet", unit_fmt)

        # Optional page setup
        ws.set_landscape()
        ws.fit_to_pages(1, 1)

        workbook.close()

    return summary_data, write_summary_sheet


@app.cell
def _(summary_data, write_summary_sheet):
    write_summary_sheet("shore_crane_report.xlsx", summary_data)
    return


@app.cell
def _():
    return


@app.cell(column=10, hide_code=True)
def _():
    mo.md(r"""
    # Openpyxl Styling for Excel
    """)
    return


@app.cell
def _():
    from typing import Iterable

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import TableStyleInfo
    from openpyxl.worksheet.properties import PageSetupProperties

    return (
        Alignment,
        Border,
        Font,
        Iterable,
        PageSetupProperties,
        PatternFill,
        Side,
        TableStyleInfo,
        get_column_letter,
        load_workbook,
    )


@app.cell
def _(dataclass):
    @dataclass
    class ReportMeta:
        report_name: str
        vessel: str
        start_date: date
        end_date: date

    return (ReportMeta,)


@app.cell
def _(
    Alignment,
    Border,
    Font,
    Iterable,
    PageSetupProperties,
    PatternFill,
    ReportMeta,
    Side,
    TableStyleInfo,
    get_column_letter,
    load_workbook,
):
    def _norm(value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_")


    def _build_header_map(ws) -> dict[str, int]:
        return {
            _norm(ws.cell(row=1, column=c).value): c
            for c in range(1, ws.max_column + 1)
        }


    def _resolve_existing(
        headers_map: dict[str, int], names: Iterable[str]
    ) -> list[str]:
        out = []
        for name in names:
            n = _norm(name)
            if n in headers_map:
                out.append(n)
        return out


    def _guess_column_roles(headers_map: dict[str, int]) -> dict[str, set[str]]:
        headers = set(headers_map.keys())

        date_like = {
            h
            for h in headers
            if any(token in h for token in ["date", "service_date"])
        }
        time_like = {
            h for h in headers if "time" in h or h in {"start_time", "end_time"}
        }
        currency_like = {
            h
            for h in headers
            if any(
                token in h
                for token in ["price", "rate", "amount", "cost", "charge"]
            )
        }
        integer_like = {
            h
            for h in headers
            if any(
                token in h
                for token in ["hours", "qty", "quantity", "count", "moves"]
            )
        }

        center_like = {
            h
            for h in headers
            if h
            in {
                "day_name",
                "service_date",
                "date",
                "start_time",
                "end_time",
                "hours",
                "overtime_hours",
                "unit",
                "shift",
            }
        }

        right_like = currency_like | {
            h
            for h in headers
            if any(token in h for token in ["tonnage", "qty", "quantity", "hours"])
        }

        return {
            "date_cols": date_like,
            "time_cols": time_like,
            "currency_cols": currency_like,
            "int_cols": integer_like,
            "center_headers": center_like,
            "right_headers": right_like,
        }


    def format_generic_report(
        xlsx_path: str,
        sheet_name: str,
        report_title: str | None = None,
        report_meta: ReportMeta | None = None,
        footer_text: str | None = None,
        table_name: str | None = None,
        header_color: str = "595959",
        currency_format: str = "#,##0.000",
        number_format_map: dict[str, str] | None = None,
        width_overrides: dict[str, float] | None = None,
        date_cols: Iterable[str] | None = None,
        time_cols: Iterable[str] | None = None,
        currency_cols: Iterable[str] | None = None,
        int_cols: Iterable[str] | None = None,
        center_headers: Iterable[str] | None = None,
        right_headers: Iterable[str] | None = None,
        auto_detect: bool = True,
        # New options
        freeze_panes: str | None = None,  # e.g. "A2" or None
        fit_to_page: bool = True,
        fit_to_width: int = 1,
        fit_to_height: int = 0,  # 0 = unlimited pages tall, 1 = single page tall
        repeat_header_row: bool = True,
        # Header/footer logo
        header_logo_text: str | None = None,  # text placeholder like company name
        header_logo_position: str = "left",  # left / center / right
        # Optional image placeholder path note:
        # true Excel header/footer image support is limited in openpyxl
        logo_path: str | None = None,
    ) -> None:
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column
        header_map = _build_header_map(ws)

        # ------------------------------------------------------------------
        # Page / worksheet layout
        # ------------------------------------------------------------------
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_view.showGridLines = False
        ws.sheet_view.view = "normal"

        if fit_to_page:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(
                fitToPage=True,
                autoPageBreaks=False,
            )
            ws.page_setup.fitToWidth = fit_to_width
            ws.page_setup.fitToHeight = fit_to_height

        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.2

        # Center horizontally when printing
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False

        # ------------------------------------------------------------------
        # Header / footer
        # ------------------------------------------------------------------

        report_meta = report_meta or {}

        if report_meta:
            report_name = report_meta.report_name or report_title
            vessel = report_meta.vessel
            start_date = report_meta.start_date
            end_date = report_meta.end_date
        else:
            report_name = report_title
            vessel = None
            start_date = None
            end_date = None

        # ---- LEFT (logo / company) ----
        if header_logo_text:
            ws.oddHeader.left.text = header_logo_text
            ws.oddHeader.left.font = "Calibri,Bold"
            ws.oddHeader.left.size = 11

        # ---- CENTER (report title) ----
        if report_name:
            ws.oddHeader.center.text = report_name
            ws.oddHeader.center.font = "Calibri,Bold"
            ws.oddHeader.center.size = 12

        # ---- RIGHT (vessel + date range) ----
        right_parts = []

        if vessel:
            right_parts.append(vessel)

        if start_date and end_date:
            right_parts.append(f"{start_date:%d-%b-%Y} to {end_date:%d-%b-%Y}")

        right_text = " | ".join(right_parts)

        if right_text:
            ws.oddHeader.right.text = right_text
            ws.oddHeader.right.font = "Calibri"
            ws.oddHeader.right.size = 10

        if footer_text:
            ws.oddFooter.center.text = footer_text
            ws.oddFooter.center.font = "Calibri"
            ws.oddFooter.center.size = 10

        ws.oddFooter.right.text = "Page &[Page] of &[Pages]"

        if repeat_header_row:
            ws.print_title_rows = "1:1"

        # ------------------------------------------------------------------
        # Optional logo / header branding
        # ------------------------------------------------------------------
        # In openpyxl, header/footer images are not reliably supported like normal worksheet images.
        # Best practical option: use text branding in header/footer.
        if header_logo_text:
            pos = header_logo_position.lower()
            if pos == "left":
                ws.oddHeader.left.text = header_logo_text
                ws.oddHeader.left.font = "Calibri,Bold"
                ws.oddHeader.left.size = 11
            elif pos == "right":
                ws.oddHeader.right.text = header_logo_text
                ws.oddHeader.right.font = "Calibri,Bold"
                ws.oddHeader.right.size = 11
            else:
                ws.oddHeader.center.text = (
                    f"{header_logo_text} | {report_title}"
                    if report_title
                    else header_logo_text
                )
                ws.oddHeader.center.font = "Calibri,Bold"
                ws.oddHeader.center.size = 12

        # NOTE:
        # logo_path is kept as a parameter for future use, but openpyxl does not
        # reliably support inserting actual images into header/footer across Excel viewers.
        # If you need a real image logo, the safer approach is to place the image in worksheet cells above the table.

        # ------------------------------------------------------------------
        # Freeze panes
        # ------------------------------------------------------------------
        ws.freeze_panes = freeze_panes

        # ------------------------------------------------------------------
        # Table styling
        # ------------------------------------------------------------------
        if ws.tables:
            if table_name and table_name in ws.tables:
                tab = ws.tables[table_name]
            else:
                first_key = next(iter(ws.tables))
                tab = ws.tables[first_key]

            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

        # ------------------------------------------------------------------
        # Header row formatting
        # ------------------------------------------------------------------
        header_fill = PatternFill("solid", fgColor=header_color)
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.row_dimensions[1].height = 24

        # ------------------------------------------------------------------
        # Detect roles
        # ------------------------------------------------------------------
        guessed = (
            _guess_column_roles(header_map)
            if auto_detect
            else {
                "date_cols": set(),
                "time_cols": set(),
                "currency_cols": set(),
                "int_cols": set(),
                "center_headers": set(),
                "right_headers": set(),
            }
        )

        date_cols_resolved = (
            set(_resolve_existing(header_map, date_cols or []))
            | guessed["date_cols"]
        )
        time_cols_resolved = (
            set(_resolve_existing(header_map, time_cols or []))
            | guessed["time_cols"]
        )
        currency_cols_resolved = (
            set(_resolve_existing(header_map, currency_cols or []))
            | guessed["currency_cols"]
        )
        int_cols_resolved = (
            set(_resolve_existing(header_map, int_cols or []))
            | guessed["int_cols"]
        )
        center_headers_resolved = (
            set(_resolve_existing(header_map, center_headers or []))
            | guessed["center_headers"]
        )
        right_headers_resolved = (
            set(_resolve_existing(header_map, right_headers or []))
            | guessed["right_headers"]
        )

        # ------------------------------------------------------------------
        # Number formats
        # ------------------------------------------------------------------
        for col_name in currency_cols_resolved:
            col_idx = header_map.get(col_name)
            if col_idx:
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = currency_format

        for col_name in int_cols_resolved:
            col_idx = header_map.get(col_name)
            if col_idx:
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = "0"

        for col_name in date_cols_resolved:
            col_idx = header_map.get(col_name)
            if col_idx:
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = "yyyy-mm-dd"

        for col_name in time_cols_resolved:
            col_idx = header_map.get(col_name)
            if col_idx:
                for r in range(2, max_row + 1):
                    ws.cell(r, col_idx).number_format = "hh:mm"

        if number_format_map:
            for col_name, fmt in number_format_map.items():
                n = _norm(col_name)
                col_idx = header_map.get(n)
                if col_idx:
                    for r in range(2, max_row + 1):
                        ws.cell(r, col_idx).number_format = fmt

        # ------------------------------------------------------------------
        # Alignment
        # ------------------------------------------------------------------
        for col_name, col_idx in header_map.items():
            if col_name in center_headers_resolved:
                align = Alignment(horizontal="center", vertical="center")
            elif col_name in right_headers_resolved:
                align = Alignment(horizontal="right", vertical="center")
            else:
                align = Alignment(horizontal="left", vertical="center")

            for r in range(2, max_row + 1):
                ws.cell(r, col_idx).alignment = align

        # ------------------------------------------------------------------
        # Borders
        # ------------------------------------------------------------------
        thin_gray = Side(style="thin", color="D9D9D9")
        bottom_border = Border(bottom=thin_gray)

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(r, c).border = bottom_border

        # ------------------------------------------------------------------
        # Auto width
        # ------------------------------------------------------------------
        width_overrides = {_norm(k): v for k, v in (width_overrides or {}).items()}

        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            header = _norm(ws.cell(1, c).value)

            if header in width_overrides:
                ws.column_dimensions[col_letter].width = width_overrides[header]
            else:
                max_len = 0
                for r in range(1, max_row + 1):
                    value = ws.cell(r, c).value
                    if value is not None:
                        max_len = max(max_len, len(str(value)))
                ws.column_dimensions[col_letter].width = min(
                    max(max_len + 2, 10), 30
                )

        # ------------------------------------------------------------------
        # Hide unused rows/columns
        # ------------------------------------------------------------------
        for r in range(max_row + 1, min(max_row + 201, 1048577)):
            ws.row_dimensions[r].hidden = True

        for c in range(max_col + 1, min(max_col + 51, 16385)):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        # ------------------------------------------------------------------
        # Print area
        # ------------------------------------------------------------------
        ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"

        wb.save(xlsx_path)

    return (format_generic_report,)


@app.cell
def _(end_date, final_shore_crane_df, start_date, vessel_name):
    summary_df = final_shore_crane_df.filter(
        (pl.col("service_date") >= start_date)
        & (pl.col("service_date") <= end_date)
        & (pl.col("customer") == vessel_name)
        & (pl.col("invoiced_to") != "MAERSKLINE")
    )
    return (summary_df,)


@app.cell(hide_code=True)
def _(summary_df):
    _df = mo.sql(
        f"""
        FROM summary_df
        """
    )
    return


@app.cell(hide_code=True)
def _(summary_df):
    _df = mo.sql(
        f"""
        WITH bucketed AS (
            SELECT
                service_date,
                day_name,
                total_hours,
                overtime_hours,
                unit_price,
                total_price,

                CASE
                    WHEN day_name NOT IN ('Sun', 'PH')
                        THEN total_hours - overtime_hours
                    ELSE 0
                END AS normal_hour_hours,

                CASE
                    WHEN day_name IN ('Sun', 'PH')
                        THEN total_hours - overtime_hours
                    WHEN day_name NOT IN ('Sun', 'PH') AND overtime_hours > 0
                        THEN overtime_hours
                    ELSE 0
                END AS overtime_150_hours,

                CASE
                    WHEN day_name IN ('Sun', 'PH') AND overtime_hours > 0
                        THEN overtime_hours
                    ELSE 0
                END AS overtime_200_hours
            FROM summary_df
        ),
        unpivoted AS (
            SELECT
                'normal_hour' AS bucket,
                normal_hour_hours AS hours,
                normal_hour_hours * unit_price AS price
            FROM bucketed

            UNION ALL

            SELECT
                'overtime_150' AS bucket,
                overtime_150_hours AS hours,
                overtime_150_hours * unit_price AS price
            FROM bucketed

            UNION ALL

            SELECT
                'overtime_200' AS bucket,
                overtime_200_hours AS hours,
                CASE
                    WHEN service_date >= DATE '2026-03-01'
                        THEN overtime_200_hours * unit_price * (2.1 / 1.6)
                    ELSE overtime_200_hours * unit_price * (2.0 / 1.5)
                END AS price
            FROM bucketed
        ),
        summary AS (
            SELECT
                bucket,
                SUM(hours) AS hours,
                ROUND(SUM(price), 3) AS price
            FROM unpivoted
            GROUP BY bucket
        )
        SELECT * FROM summary

        UNION ALL

        SELECT
            'total' AS bucket,
            SUM(hours) AS hours,
            ROUND(SUM(price), 3) AS price
        FROM summary;
        """
    )
    return


@app.cell
def _():
    return


@app.cell(column=11, hide_code=True)
def _():
    mo.md(r"""
    ## Data to process to Excel
    """)
    return


@app.cell
def _(ReportMeta):
    start_date = date(year=2026, month=2, day=23)
    end_date = date(2026, 2, 26)
    vessel_name = "TORRE ITALIA"


    meta = ReportMeta(
        report_name="STO",
        vessel=vessel_name,
        start_date=start_date,
        end_date=end_date,
    )
    return end_date, meta, start_date, vessel_name


@app.cell(column=12, hide_code=True)
def _():
    mo.md(r"""
    # Daily Unloading ✔️
    """)
    return


@app.cell(hide_code=True)
def _(client_raw_dataf):
    vessel_list = mo.sql(
        f"""
        FROM
            client_raw_dataf
        SELECT
            "Vessel/Client" AS vessel,
            -- "Customer" AS ship_owner
        WHERE
            "Type" IN (
                '{VesselType.PURSEINER}')
        """,
        output=False,
    )
    return (vessel_list,)


@app.cell(hide_code=True)
def _(date_select, final_net_list_df, vessel_select):
    time_df = mo.sql(
        f"""
        FROM final_net_list_df
            SELECT MIN(start_time) AS start_time,
            MAX(end_time) AS end_time

        WHERE vessel = '{vessel_select.value}' AND service_date = '{date_select.value}'
        GROUP BY ALL
        """,
        output=False,
    )
    return (time_df,)


@app.cell(hide_code=True)
def _(operations):
    genesis_df = scan_google_sheet(
        url=operations.url, sheet_name=operations.sheet("genesis")
    )

    well_to_well_df = scan_google_sheet(
        url=operations.url, sheet_name=operations.sheet("well_to_well")
    )
    return genesis_df, well_to_well_df


@app.cell(hide_code=True)
def _(date_select, vessel_select, well_to_well_df):
    well_df = mo.sql(
        f"""
        FROM well_to_well_df
            SELECT Tonnage
        WHERE Vessel = '{vessel_select.value}' AND Date = '{date_select.value}'
        """,
        output=False
    )
    return (well_df,)


@app.cell(hide_code=True)
def _(date_select, genesis_df, vessel_select):
    side_working = mo.sql(
        f"""
        FROM genesis_df
            SELECT STRING_AGG(DISTINCT "Side Working",',') AS side_worked
        WHERE UPPER(Vessel) = '{vessel_select.value}' AND Date = '{date_select.value}'
        """,
        output=False,
    )
    return (side_working,)


@app.cell(hide_code=True)
def _(date_select, final_net_list_df, vessel_select):
    _df = mo.sql(
        f"""
        FROM final_net_list_df
        WHERE vessel = '{vessel_select.value}' AND service_date = '{date_select.value}'
        """
    )
    return


@app.cell(hide_code=True)
def _(date_select, final_net_list_df, vessel_select):
    summary_unloading = mo.sql(
        f"""
        FROM final_net_list_df
            SELECT day_name,
            -- destination,
            overtime,
            storage_type,
            CAST(SUM(total_tonnage) AS DECIMAL) AS tonnage

        WHERE vessel = '{vessel_select.value}' AND service_date = '{date_select.value}'

        GROUP BY ALL
        """,
        output=False,
    )
    return (summary_unloading,)


@app.cell(hide_code=True)
def _(date_select, final_net_list_df, vessel_select):
    overtime_table = mo.sql(
        f"""
        WITH add_ot AS (FROM final_net_list_df
            SELECT *,CASE WHEN day_name IN ('PH','Sun') AND overtime = 'overtime 200%' THEN 'OT' WHEN day_name NOT IN ('Sun','PH') AND overtime = 'overtime 150%' THEN 'OT' ELSE NULL END AS checked_ot
        WHERE vessel = '{vessel_select.value}' AND service_date = '{date_select.value}')

        FROM add_ot
            SELECT service,overtime,storage_type,MAX(end_time) AS end_time,ROUND(SUM(total_tonnage),3) AS overtime_tonnage
        WHERE checked_ot = 'OT'
        GROUP BY ALL
        """
    )
    return (overtime_table,)


@app.cell(hide_code=True)
def _(side_working, summary_unloading, time_df, well_df):
    start = mo.md(
        f"Start: {time_df.get_column('start_time').first().strftime(format='%H:%M')}"
    )
    end = mo.md(
        f"End: {time_df.get_column('end_time').first().strftime(format='%H:%M')}"
    )
    side = mo.md(f"Sides: {side_working.get_column('side_worked').first()}")

    well = mo.md(f"Well to Well: {well_df.get_column('Tonnage').first()} tons")
    total = mo.md(
        f"Total: {summary_unloading.select(pl.col('tonnage').cast(pl.Float64)).sum().to_series().to_list()[0]:,.3f}"
    )
    return end, side, start, total, well


@app.cell(hide_code=True)
def _(vessel_list):
    vessel_select = mo.ui.dropdown(
        options=vessel_list.to_series(), label="Vessel: ", searchable=True
    )
    date_select = mo.ui.date(label="Date: ")

    mo.vstack([vessel_select, date_select])
    return date_select, vessel_select


@app.cell(hide_code=True)
def _(
    date_select,
    end,
    overtime_table,
    side,
    start,
    summary_unloading,
    total,
    vessel_select,
    well,
):
    mo.vstack(
        [
            mo.hstack([vessel_select.value, date_select.value]),
            start,
            end,
            mo.md("---"),
            side,
            mo.ui.table(summary_unloading),
            well,
            total,
            mo.md("---"),
            mo.md("overtime."),
            overtime_table,
        ]
    )
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
