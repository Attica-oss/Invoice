import marimo

__generated_with = "0.23.1"
app = marimo.App(width="medium")

with app.setup:
    import polars as pl
    from openpyxl import load_workbook
    import marimo as mo
    from pathlib import Path
    from scan_google_sheet import scan_google_sheet


@app.cell
def _():
    BASE_PATH = Path.home() / "Dropbox"

    OPERATIONS_ACTIVITY_2026 = (
        Path(
            rf"""{BASE_PATH}/! OPERATION SUPPORTING DOCUMENTATION/2026/2026 IPHS operation activity.xlsx"""
        ),
        "HANDLING ACTIVITY",
    )
    return (OPERATIONS_ACTIVITY_2026,)


@app.cell
def _(OPERATIONS_ACTIVITY_2026):
    ops_df = pl.read_excel(source=OPERATIONS_ACTIVITY_2026[0],sheet_name=OPERATIONS_ACTIVITY_2026[1]).filter(pl.col("DATE").is_not_null())
    return (ops_df,)


@app.cell
def _():
    well_df = scan_google_sheet(url="https://docs.google.com/spreadsheets/d/1PvTkl6DYZdhtaiNshz0qwtSPxC8S1OOeu905NmhFKNs/edit?gid=1301483182#gid=1301483182",sheet_name="WelltoWell").filter(pl.col("Date").dt.year().eq(2026))
    return (well_df,)


@app.cell(hide_code=True)
def _(ops_df, well_df):
    _df = mo.sql(
        f"""
        WITH well_ops AS (FROM
            ops_df
        SELECT
            "DAY" AS day_name,
            "DATE" AS date,
            "VESSEL NAME" AS vessel,
            "Well-to-Well Transfer" AS well_to_well
        WHERE
            "Well-to-Well Transfer" IS NOT NULL),well_inv AS (FROM well_df)

        FROM well_ops o FULL OUTER JOIN well_inv i ON i.Date = o.date AND i.Vessel = o.vessel
        ORDER BY o.date
        """
    )
    return


@app.cell
def _():
    extra_inv_df = scan_google_sheet(url="https://docs.google.com/spreadsheets/d/1PvTkl6DYZdhtaiNshz0qwtSPxC8S1OOeu905NmhFKNs/edit?gid=1301483182#gid=1301483182",sheet_name="ExtraMen")
    return (extra_inv_df,)


@app.cell(hide_code=True)
def _(extra_inv_df, ops_df):
    _df = mo.sql(
        f"""
        WITH
            extra_ops AS (
                FROM
                    ops_df
                SELECT
                    "DAY" AS day_name,
                    "DATE" AS date,
                    "VESSEL NAME" AS vessel,
                    CAST("TOTAL TONNAGE" AS DECIMAL) AS tonnage,
                    "Extra Men" AS extra_men,
                    "Number of Stevedores" AS num_of_stevedores,
                    "Comments" AS remarks
            ),
            extra_inv AS (
                FROM
                    extra_inv_df
                SELECT
                    Day,
                    Date,
                    Vessel,
                    ExtraMen,
                    CAST(NULLIF(TRIM(NumberOfStevedores), '') AS BIGINT) AS NumberOfStevedores,
                    "Comments" AS remarks
            )
    
        SELECT
            o.*,
            i.ExtraMen,
            i.NumberOfStevedores
        FROM
            extra_ops o
            FULL OUTER JOIN extra_inv i ON o.day_name = i.Day
            AND i.Date = o.date
            AND i.Vessel = o.vessel
        WHERE
            MONTH(o.date) = 4
            -- AND (
            --     i.NumberOfStevedores IS NULL
            -- )
        ORDER BY
            o.date,o.vessel

        """
    )
    return


@app.cell
def _(OPERATIONS_ACTIVITY_2026):
    pl.read_excel(
        source=OPERATIONS_ACTIVITY_2026[0],
        sheet_name="Additional Stevedores",
        schema_overrides={"End Time": pl.Time,"Tonnage":pl.Decimal(precision=5,scale=3)},
    ).filter(pl.col("Tonnage").is_not_null()).select(
        pl.all().exclude(["Price", "Additional Stevedores ($)"])
    )
    return


@app.cell
def _(OPERATIONS_ACTIVITY_2026):
    pl.read_excel(source=OPERATIONS_ACTIVITY_2026[0],sheet_name="Extramen").filter(~pl.col("Check"))
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
